"""
Assegnazione automatica OEPAC — vincolo 45h DGC 260/2024

Strumento per l'assegnazione degli alunni con disabilità (servizio OEPAC)
agli Organismi accreditati, secondo le Linee Guida approvate con
DGC Roma Capitale n. 260/2024 (Art. 5, commi 5 e 6).

Classificazione automatica degli ambiti (Roma Capitale): i gruppi 45h sono
derivati automaticamente per Istituto Comprensivo statale e, separatamente,
per le scuole dell'infanzia comunali e per le scuole paritarie (per ambito
territoriale — numerazione cittadina 1..37 — o per municipio). Supporta le
assegnazioni in corso d'anno secondo le quattro finestre di attivazione
(nota Dipartimento Scuola QM/102670/2025).
"""

import datetime
import html
import io
import re

import pandas as pd
import streamlit as st


# ---------------------------------------------------------------------------
# Utility functions
# ---------------------------------------------------------------------------

def normalizza_nome(nome: str) -> str:
    if not isinstance(nome, str):
        return ""
    return re.sub(r"\s+", " ", nome.strip()).casefold()


def nomi_uguali(a: str, b: str) -> bool:
    return normalizza_nome(a) == normalizza_nome(b)



def parse_data_nascita(val) -> datetime.date | None:
    if isinstance(val, datetime.datetime):
        return val.date()
    if isinstance(val, datetime.date):
        return val
    if pd.isna(val):
        return None
    s = str(val).strip().split(" ")[0]
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d.%m.%Y"):
        try:
            return datetime.datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def calcola_eta(data_nascita: datetime.date | None) -> int | None:
    if data_nascita is None:
        return None
    oggi = datetime.date.today()
    eta = oggi.year - data_nascita.year
    if (oggi.month, oggi.day) < (data_nascita.month, data_nascita.day):
        eta -= 1
    return eta


def deriva_grado(codice_mecc_plesso: str, contesto: str = "") -> str:
    """Deriva il grado scolastico dal codice meccanografico del plesso.

    `contesto` è un testo libero (Tipo Gestione derivato oppure la colonna
    Ambito del MESIS) usato per distinguere le infanzie comunali dalle
    paritarie, che condividono il prefisso 1A.
    """
    if not isinstance(codice_mecc_plesso, str) or len(codice_mecc_plesso) < 4:
        return "N/D"
    prefix = codice_mecc_plesso[2:4].upper()
    contesto_up = contesto.upper().strip() if isinstance(contesto, str) else ""
    is_paritario = "PARIT" in contesto_up

    if prefix == "AA":
        return "Infanzia statale"
    if prefix == "EE":
        return "Primaria statale"
    if prefix == "MM":
        return "Sec. I grado statale"
    if prefix == "1A":
        if is_paritario:
            return "Infanzia paritaria"
        return "Infanzia comunale"
    if prefix == "1E":
        return "Primaria paritaria"
    if prefix == "1M":
        return "Sec. I grado paritaria"
    return "N/D"


# ---------------------------------------------------------------------------
# Classificazione automatica ambiti (Roma Capitale)
#
# Deriva da codice meccanografico, colonna Municipio e denominazioni:
#   - Statale          -> gruppo 45h per Istituto Comprensivo (IC:<codice>)
#   - Infanzia comunale-> gruppo 45h per municipio (COM:Municipio <n>)
#   - Paritaria        -> gruppo 45h per municipio (PAR:Municipio <n>)
# ---------------------------------------------------------------------------

MUNICIPI_ROMA = [
    "I", "II", "III", "IV", "V", "VI", "VII", "VIII",
    "IX", "X", "XI", "XII", "XIII", "XIV", "XV",
]
_ROMANO_A_NUMERO = {r: i + 1 for i, r in enumerate(MUNICIPI_ROMA)}

PREFISSI_PLESSO_STATALE = {"AA", "EE", "MM"}
KEYWORDS_COMUNALE = ("COMUNAL", "CAPITOLIN")

TIPO_STATALE = "Statale"
TIPO_COMUNALE = "Infanzia comunale"
TIPO_PARITARIA = "Paritaria"
TIPI_GESTIONE = [TIPO_STATALE, TIPO_COMUNALE, TIPO_PARITARIA]


def normalizza_municipio(val) -> str:
    """Normalizza il municipio di Roma in numero romano ('I'..'XV').

    Accetta '1', '01', 'VII', 'Municipio 7', 'MUNICIPIO XIII', 7, ecc.
    Ritorna stringa vuota se non riconosciuto.
    """
    if val is None:
        return ""
    try:
        if pd.isna(val):
            return ""
    except (TypeError, ValueError):
        pass
    s = str(val).strip().upper()
    if not s:
        return ""
    n = None
    m = re.search(r"\b(\d{1,2})\b", s)
    if m:
        n = int(m.group(1))
    else:
        m = re.search(r"\b([IVX]{1,4})\b", s)
        if m and m.group(1) in _ROMANO_A_NUMERO:
            n = _ROMANO_A_NUMERO[m.group(1)]
    if n is not None and 1 <= n <= len(MUNICIPI_ROMA):
        return MUNICIPI_ROMA[n - 1]
    return ""


def estrai_numero_ambito(ambito: str) -> str:
    """Estrae il numero dell'ambito territoriale dalla colonna Ambito MESIS.

    Esempi: '10' -> '10', '12 - Paritario' -> '12', 'IC' -> ''.
    """
    if not isinstance(ambito, str):
        return ""
    m = re.search(r"\d+", ambito)
    return m.group(0) if m else ""


def classifica_gestione(
    codice_plesso: str,
    ambito: str = "",
    istituto: str = "",
    plesso: str = "",
    codici_comunali: set[str] | None = None,
) -> tuple[str, str]:
    """Classifica la scuola come Statale / Infanzia comunale / Paritaria.

    Ritorna (tipo, fonte) dove fonte indica come è stata determinata la
    classificazione: 'elenco comunali', 'codice', 'testo', 'ambito MESIS',
    'presunta' o ''.
    """
    cod = codice_plesso.strip().upper() if isinstance(codice_plesso, str) else ""
    if codici_comunali and cod in codici_comunali:
        return TIPO_COMUNALE, "elenco comunali"
    if len(cod) < 4:
        return "N/D", ""
    prefix = cod[2:4]
    if prefix in PREFISSI_PLESSO_STATALE:
        return TIPO_STATALE, "codice"
    if prefix[0] != "1":
        return "N/D", ""
    if prefix in ("1E", "1M"):
        # primarie e secondarie non statali: a Roma sono paritarie
        # (Roma Capitale gestisce direttamente solo scuole dell'infanzia)
        return TIPO_PARITARIA, "codice"
    # prefisso 1A (o altro non statale): infanzia comunale o paritaria
    testo = " ".join(
        v.upper() for v in (ambito, istituto, plesso) if isinstance(v, str)
    )
    if "PARIT" in testo:
        return TIPO_PARITARIA, "testo"
    if any(k in testo for k in KEYWORDS_COMUNALE):
        return TIPO_COMUNALE, "testo"
    # nel MESIS le infanzie comunali stanno negli ambiti territoriali
    # numerici (es. '10', '11'), le paritarie in quelli '... - Paritario'
    if estrai_numero_ambito(ambito):
        return TIPO_COMUNALE, "ambito MESIS"
    # A Roma le infanzie non statali del circuito OEPAC sono in prevalenza
    # comunali: in assenza di altri segnali la classificazione va verificata.
    return TIPO_COMUNALE, "presunta"


def deriva_gruppo_auto(
    tipo: str,
    municipio: str,
    codice_istituto: str,
    codice_plesso: str,
    ambito_orig: str = "",
) -> str:
    """Deriva il gruppo 45h dalla classificazione automatica.

    Statali: un gruppo per Istituto Comprensivo. Comunali e paritarie:
    l'ambito territoriale numerico del MESIS se presente (la numerazione
    degli ambiti è cittadina, 1..37, quindi già univoca), altrimenti il
    municipio; i due circuiti restano comunque separati (COM:/PAR:).
    """
    if tipo == TIPO_STATALE:
        if codice_istituto:
            return f"IC:{codice_istituto}"
        return f"PLESSO:{codice_plesso}"
    n_amb = estrai_numero_ambito(ambito_orig)
    if tipo == TIPO_COMUNALE:
        if n_amb:
            return f"COM:Ambito {n_amb}"
        return f"COM:Municipio {municipio or 'N/D'}"
    if tipo == TIPO_PARITARIA:
        if n_amb:
            return f"PAR:Ambito {n_amb}"
        return f"PAR:Municipio {municipio or 'N/D'}"
    if ambito_orig:
        return f"AMB:{ambito_orig}"
    return f"PLESSO:{codice_plesso}"


def etichetta_gruppo(codice_gruppo: str, descrizione: str = "") -> str:
    """Etichetta leggibile del gruppo 45h per i documenti.

    Es. 'IC:RMIC8E0001' -> 'Istituto Comprensivo (RMIC8E0001)',
    'COM:Ambito 10' -> 'Infanzia comunale — Ambito 10'.
    """
    if not isinstance(codice_gruppo, str) or ":" not in codice_gruppo:
        return codice_gruppo or ""
    prefisso, resto = codice_gruppo.split(":", 1)
    mappa = {
        "IC": "Istituto Comprensivo",
        "COM": "Infanzia comunale",
        "PAR": "Scuola paritaria",
        "AMB": "Ambito",
        "PLESSO": "Plesso",
    }
    base = mappa.get(prefisso, prefisso)
    if prefisso == "IC":
        etichetta = f"{base} {descrizione}".strip() if descrizione else base
        return f"{etichetta} ({resto})" if descrizione else f"{base} {resto}"
    if prefisso in ("COM", "PAR"):
        return f"{base} — {resto}"
    return f"{base} {resto}"


def _primo_non_vuoto(valori) -> str:
    for v in valori:
        if isinstance(v, str) and v.strip():
            return v.strip()
    return ""


def risolvi_plessi(
    df: pd.DataFrame,
    col_plesso: str,
    col_ambito: str | None = None,
    col_istituto_nome: str | None = None,
    col_plesso_nome: str | None = None,
    col_municipio: str | None = None,
    col_istituto_cod: str | None = None,
    codici_comunali: set[str] | None = None,
    tipi_override: dict[str, str] | None = None,
) -> dict[str, dict[str, str]]:
    """Risolve classificazione, municipio e ambito UNA volta per plesso.

    Tutte le righe di uno stesso plesso ricevono così la stessa
    classificazione e lo stesso gruppo 45h anche quando i valori MESIS
    (Ambito, Municipio, denominazioni) variano o mancano su alcune righe.
    La classificazione considera tutti i valori Ambito del plesso, così un
    solo '– Paritario' basta a qualificarlo.
    """
    override_norm = {
        str(k).strip().upper(): v for k, v in (tipi_override or {}).items()
    }

    def _serie(col):
        if col and col in df.columns:
            return df[col]
        return pd.Series("", index=df.index)

    ambiti = _serie(col_ambito)
    municipi_norm = _serie(col_municipio).apply(normalizza_municipio)
    nomi_ist = _serie(col_istituto_nome)
    nomi_plesso = _serie(col_plesso_nome)
    codici_ist = _serie(col_istituto_cod)

    risoluzione: dict[str, dict[str, str]] = {}
    for cod, indici in df.groupby(df[col_plesso], sort=True).groups.items():
        cod = str(cod).strip()
        if not cod:
            continue
        amb_valori = [
            a.strip() for a in ambiti.loc[indici]
            if isinstance(a, str) and a.strip()
        ]
        ambito = amb_valori[0] if amb_valori else ""
        ambito_testo = " ".join(sorted(set(amb_valori)))
        municipio = _primo_non_vuoto(municipi_norm.loc[indici])
        nome_ist = _primo_non_vuoto(nomi_ist.loc[indici])
        nome_ple = _primo_non_vuoto(nomi_plesso.loc[indici])
        cod_ist = _primo_non_vuoto(codici_ist.loc[indici])
        if cod.upper() in override_norm:
            tipo, fonte = override_norm[cod.upper()], "manuale"
        else:
            tipo, fonte = classifica_gestione(
                cod, ambito=ambito_testo, istituto=nome_ist,
                plesso=nome_ple, codici_comunali=codici_comunali,
            )
        risoluzione[cod] = {
            "tipo": tipo,
            "fonte": fonte,
            "municipio": municipio,
            "ambito": ambito,
            "codice_istituto": cod_ist,
            "gruppo": deriva_gruppo_auto(tipo, municipio, cod_ist, cod, ambito),
            "istituto": nome_ist,
            "plesso": nome_ple,
        }
    return risoluzione


def carica_codici_comunali(file_bytes: bytes, filename: str = "") -> set[str]:
    """Carica un elenco di codici meccanografici di scuole comunali.

    Accetta .xlsx/.csv con una colonna il cui nome contiene 'codice' o 'mecc'
    (in mancanza usa la prima colonna).
    """
    try:
        if filename.lower().endswith(".csv"):
            df = pd.read_csv(io.BytesIO(file_bytes), dtype=str, sep=None, engine="python")
        else:
            try:
                df = pd.read_excel(io.BytesIO(file_bytes), dtype=str)
            except Exception:
                df = pd.read_csv(io.BytesIO(file_bytes), dtype=str, sep=None, engine="python")
        target = None
        for c in df.columns:
            cl = str(c).strip().lower()
            if "codice" in cl or "mecc" in cl:
                target = c
                break
        codes: set[str] = set()
        if target is None:
            if len(df.columns) != 1:
                return set()
            target = df.columns[0]
            # file a colonna singola senza intestazione: anche il nome
            # colonna potrebbe essere un codice
            header = str(target).strip().upper()
            if re.fullmatch(r"[A-Z]{2}\w{8}", header):
                codes.add(header)
        codes.update(
            str(v).strip().upper() for v in df[target].dropna() if str(v).strip()
        )
        return codes
    except Exception:
        return set()


def parse_preferenze(testo: str) -> list[str]:
    if not isinstance(testo, str) or not testo.strip():
        return []
    righe = testo.split("\n")
    preferenze = []
    for riga in righe:
        riga = riga.strip()
        if not riga:
            continue
        if re.match(r"^Domanda\s+\d+\s*:", riga):
            continue
        preferenze.append(riga)
    return preferenze


# ---------------------------------------------------------------------------
# Column mapping — maps spec names to actual Excel header variations
# ---------------------------------------------------------------------------

COLUMN_MAP = [
    ("codice_iscrizione", ["Codice"]),
    ("stato", ["Stato"]),
    ("tipo", ["Tipo"]),
    ("municipio", ["Municipio"]),
    ("cognome", ["Cognome"]),
    ("nome", ["Nome"]),
    ("data_nascita", ["Data Nascita", "Data di Nascita"]),
    ("codice_mecc_istituto", ["Codice Meccanografico Istituto"]),
    ("istituto", ["Istituto"]),
    ("codice_mecc_plesso", ["Codice Meccanografico Plesso"]),
    ("plesso", ["Plesso"]),
    ("classe", ["Classe"]),
    ("sezione", ["Sezione"]),
    ("ambito", ["Ambito"]),
    ("ore_richieste", ["Ore Richieste"]),
    ("ore_assegnate", ["Ore Assegnate"]),
    ("codice_utente", ["UTENTE"]),
    ("organismo_assegnato", ["Organismo Assegnato"]),
    ("enti_erogatori_scelti", ["Enti Erogatori scelti", "Enti Erogatori Scelti"]),
    ("data_attivazione", ["Data Attivazione"]),
    ("data_rinuncia", ["Data Rinuncia/Sospensione", "Data Rinuncia"]),
]


def map_columns(df: pd.DataFrame) -> dict[str, str]:
    headers = list(df.columns)
    headers_lower = [h.strip().lower() if isinstance(h, str) else "" for h in headers]
    used: set[int] = set()
    mapping: dict[str, str] = {}

    for key, candidates in COLUMN_MAP:
        for cand in candidates:
            cl = cand.strip().lower()
            for i, h in enumerate(headers_lower):
                if i in used:
                    continue
                if cl == h:
                    mapping[key] = headers[i]
                    used.add(i)
                    break
            if key in mapping:
                break

    for key, candidates in COLUMN_MAP:
        if key in mapping:
            continue
        for cand in candidates:
            cl = cand.strip().lower()
            for i, h in enumerate(headers_lower):
                if i in used:
                    continue
                if cl in h:
                    mapping[key] = headers[i]
                    used.add(i)
                    break
            if key in mapping:
                break

    return mapping


# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------

def carica_dati(file_bytes: bytes) -> tuple[pd.DataFrame, dict[str, str], list[str]]:
    errors = []
    df_raw = pd.read_excel(
        io.BytesIO(file_bytes),
        sheet_name=0,
        header=None,
        dtype=str,
    )

    header_row_idx = 7
    if df_raw.shape[0] <= header_row_idx:
        errors.append(
            f"Il file ha solo {df_raw.shape[0]} righe — mi aspetto almeno 9 "
            "(7 di metadati + 1 intestazione + dati)."
        )
        return pd.DataFrame(), {}, errors

    raw_headers = [
        re.sub(r"\s+", " ", str(v).strip()) if pd.notna(v) else f"Col_{i}"
        for i, v in enumerate(df_raw.iloc[header_row_idx])
    ]
    seen: dict[str, int] = {}
    headers = []
    for h in raw_headers:
        if h in seen:
            seen[h] += 1
            headers.append(f"{h}_{seen[h]}")
        else:
            seen[h] = 0
            headers.append(h)

    df = df_raw.iloc[header_row_idx + 1 :].copy()
    df.columns = headers
    df.reset_index(drop=True, inplace=True)

    col_map = map_columns(df)

    required = [
        "codice_iscrizione", "stato", "tipo", "cognome", "nome",
        "codice_mecc_plesso", "ore_assegnate",
    ]
    col_map_lookup = dict(COLUMN_MAP)
    for req in required:
        if req not in col_map:
            errors.append(
                f"Colonna obbligatoria non trovata: {req} "
                f"(cercate: {col_map_lookup.get(req, [])})"
            )

    return df, col_map, errors


# ---------------------------------------------------------------------------
# Algorithm
# ---------------------------------------------------------------------------

def esegui_assegnazione(
    df: pd.DataFrame,
    col_map: dict[str, str],
    soglia_ore: int = 45,
    max_iter: int = 50,
    auto_ambito: bool = False,
    tipi_override: dict[str, str] | None = None,
    codici_comunali: set[str] | None = None,
    corso_anno: bool = False,
    data_riferimento: datetime.date | None = None,
):

    log = []
    stats = {
        "totale_alunni": 0,
        "riconferme": 0,
        "nuove_iscrizioni": 0,
        "gia_attivati": 0,
        "criticita": 0,
        "spostamenti": 0,
        "iterazioni": 0,
    }

    def col(key):
        return col_map.get(key)

    stato_col = col("stato")
    rinuncia_col = col("data_rinuncia")
    tipo_col = col("tipo")
    plesso_col = col("codice_mecc_plesso")
    ore_col = col("ore_assegnate")
    org_col = col("organismo_assegnato")
    enti_col = col("enti_erogatori_scelti")
    codice_col = col("codice_iscrizione")

    mask = df[stato_col].str.strip().str.upper() == "ATTIVA"
    if rinuncia_col:
        mask &= df[rinuncia_col].isna() | (df[rinuncia_col].str.strip() == "")
    df_work = df.loc[mask].copy()
    df_work.reset_index(drop=True, inplace=True)

    empty_result_cols = [
        "Codice Iscrizione", "Tipo", "Cognome", "Nome", "Data Nascita",
        "Età", "Codice Meccanografico Istituto", "Istituto",
        "Codice Meccanografico Plesso", "Plesso", "Indirizzo Plesso",
        "Grado Scolastico", "Classe", "Sezione", "Ambito",
        "Tipo Gestione", "Fonte Classificazione", "Municipio", "Gruppo 45h",
        "Ore Richieste", "Ore Assegnate", "Organismo Pre-esistente",
        "Organismo Assegnato dall'Algoritmo", "Preferenza Soddisfatta",
        "Data Attivazione", "Status", "Note",
    ]
    if len(df_work) == 0:
        df_empty = pd.DataFrame(columns=empty_result_cols)
        df_riep_empty = pd.DataFrame(columns=[
            "Gruppo 45h", "Tipo gruppo", "Descrizione", "Ambito", "Municipio",
            "N. plessi", "Organismo", "N. alunni assegnati",
            "Ore totali settimanali org.", "Ore totali del gruppo", "Soglia 45h",
        ])
        df_crit_empty = pd.DataFrame(columns=empty_result_cols + ["Azione suggerita"])
        log.append("Nessun alunno con Stato=ATTIVA trovato.")
        return df_empty, df_riep_empty, df_crit_empty, log, stats, {}

    ambito_col = col("ambito")
    ist_col = col("codice_mecc_istituto")
    municipio_src_col = col("municipio")
    istituto_nome_col = col("istituto")
    plesso_nome_col_src = col("plesso")

    df_work["_ore"] = pd.to_numeric(df_work[ore_col], errors="coerce").fillna(0)
    df_work["_tipo_norm"] = df_work[tipo_col].str.strip().str.upper().fillna("")
    df_work["_plesso"] = df_work[plesso_col].str.strip().fillna("")
    df_work["_org_orig"] = df_work[org_col].str.strip().fillna("") if org_col else ""
    df_work["_codice"] = df_work[codice_col].str.strip().fillna("")
    df_work["_ambito"] = df_work[ambito_col].str.strip().fillna("") if ambito_col else ""
    df_work["_istituto"] = df_work[ist_col].str.strip().fillna("") if ist_col else ""
    df_work["_municipio"] = (
        df_work[municipio_src_col].apply(normalizza_municipio)
        if municipio_src_col else ""
    )
    df_work["_nome_ist"] = (
        df_work[istituto_nome_col].fillna("") if istituto_nome_col else ""
    )
    df_work["_nome_plesso"] = (
        df_work[plesso_nome_col_src].fillna("") if plesso_nome_col_src else ""
    )

    # classificazione, municipio e ambito risolti una volta per plesso:
    # tutte le righe dello stesso plesso finiscono nello stesso gruppo
    # anche se i valori MESIS variano o mancano su singole righe
    risoluzione = risolvi_plessi(
        df_work,
        col_plesso="_plesso",
        col_ambito="_ambito",
        col_istituto_nome="_nome_ist",
        col_plesso_nome="_nome_plesso",
        col_municipio="_municipio",
        col_istituto_cod="_istituto",
        codici_comunali=codici_comunali,
        tipi_override=tipi_override,
    )

    def _campo(row, campo, default=""):
        info = risoluzione.get(row["_plesso"])
        return info[campo] if info else default

    df_work["_tipo_gest"] = df_work.apply(lambda r: _campo(r, "tipo", "N/D"), axis=1)
    df_work["_fonte_class"] = df_work.apply(lambda r: _campo(r, "fonte"), axis=1)
    df_work["_municipio"] = df_work.apply(
        lambda r: _campo(r, "municipio", r["_municipio"]), axis=1,
    )

    def calcola_gruppo_45h(row) -> str:
        ambito = row["_ambito"]
        if ambito.upper() == "IC":
            return f"IC:{row['_istituto']}"
        if ambito:
            return f"AMB:{ambito}"
        return f"PLESSO:{row['_plesso']}"

    if auto_ambito:
        df_work["_gruppo"] = df_work.apply(
            lambda r: _campo(r, "gruppo", f"PLESSO:{r['_plesso']}"), axis=1,
        )
    else:
        df_work["_gruppo"] = df_work.apply(calcola_gruppo_45h, axis=1)

    nome_registro: dict[str, str] = {}

    def registra_nome(nome: str):
        if isinstance(nome, str) and nome.strip():
            nome_registro[normalizza_nome(nome)] = nome.strip()

    for _, row in df_work.iterrows():
        if isinstance(row.get("_org_orig"), str) and row["_org_orig"]:
            registra_nome(row["_org_orig"])
        if enti_col and isinstance(row.get(enti_col), str):
            for pref in parse_preferenze(row[enti_col]):
                registra_nome(pref)

    df_work["_preferenze"] = df_work[enti_col].apply(parse_preferenze) if enti_col else [[] for _ in range(len(df_work))]

    is_riconferma = df_work["_tipo_norm"] == "RICONFERMA"
    is_nuova = df_work["_tipo_norm"] == "NUOVA ISCRIZIONE"
    is_altro = ~is_riconferma & ~is_nuova

    # In corso d'anno (finestre di attivazione) gli alunni già attivati —
    # Organismo Assegnato e Data Attivazione presenti — mantengono il loro
    # organismo per continuità: l'algoritmo assegna solo le nuove domande
    # non ancora attivate.
    data_att_col_w = col("data_attivazione")

    def _gia_attivato(idx) -> bool:
        org = df_work.at[idx, "_org_orig"]
        if not (isinstance(org, str) and org.strip()):
            return False
        if not data_att_col_w:
            return False
        val = df_work.at[idx, data_att_col_w]
        d = parse_data_nascita(val)
        if d is None:
            # data presente ma non interpretabile: prudenzialmente attivato
            return isinstance(val, str) and val.strip() != ""
        if data_riferimento is not None:
            return d <= data_riferimento
        return True

    if corso_anno:
        is_attivato = pd.Series(
            [bool(is_nuova.at[i]) and _gia_attivato(i) for i in df_work.index],
            index=df_work.index,
        )
    else:
        is_attivato = pd.Series(False, index=df_work.index)
    is_fisso = is_riconferma | is_attivato
    is_nuova_da_assegnare = is_nuova & ~is_attivato

    stats["totale_alunni"] = len(df_work)
    stats["riconferme"] = int(is_riconferma.sum())
    stats["nuove_iscrizioni"] = int(is_nuova.sum())
    stats["gia_attivati"] = int(is_attivato.sum())
    if corso_anno:
        log.append(
            f"Modalità in corso d'anno: {stats['gia_attivati']} alunni già "
            f"attivati mantenuti sul loro organismo, "
            f"{int(is_nuova_da_assegnare.sum())} nuove domande da assegnare."
        )

    df_work["_assegnato"] = ""
    df_work["_pref_idx"] = 0
    df_work["_status"] = ""
    df_work["_note"] = ""
    df_work["_pref_soddisfatta"] = ""

    # -------------------------------------------------------------------
    # Cronologia del procedimento (trasparenza / accesso agli atti):
    # registra lo stato delle assegnazioni ad ogni fase e ogni singolo
    # spostamento operato dall'algoritmo, con la relativa motivazione.
    # -------------------------------------------------------------------
    cognome_col = col("cognome")
    nome_col = col("nome")

    def _categoria(idx) -> str:
        if is_attivato.at[idx]:
            return "Già attivato"
        if is_riconferma.at[idx]:
            return "Riconferma"
        if is_nuova.at[idx]:
            return "Nuova iscrizione"
        return "Altro"

    cronologia_anagrafica: dict = {}
    for idx in df_work.index:
        cronologia_anagrafica[idx] = {
            "codice": df_work.at[idx, "_codice"],
            "cognome": df_work.at[idx, cognome_col] if cognome_col else "",
            "nome": df_work.at[idx, nome_col] if nome_col else "",
            "plesso": df_work.at[idx, "_nome_plesso"],
            "gruppo": df_work.at[idx, "_gruppo"],
            "gruppo_label": etichetta_gruppo(
                df_work.at[idx, "_gruppo"], df_work.at[idx, "_nome_ist"]
            ),
            "ore": df_work.at[idx, "_ore"],
            "categoria": _categoria(idx),
        }

    cronologia_passi: list = []
    cronologia_movimenti: list = []

    def _registra_passo(titolo, descrizione, norma="", stato=None):
        if stato is None:
            stato = {}
            for i in df_work.index:
                v = df_work.at[i, "_assegnato"]
                stato[i] = v.strip() if isinstance(v, str) and v.strip() else ""
        cronologia_passi.append({
            "indice": len(cronologia_passi),
            "titolo": titolo,
            "descrizione": descrizione,
            "norma": norma,
            "stato": stato,
        })

    def _registra_movimento(i, passo_indice, passo_titolo, da, a, motivo):
        cronologia_movimenti.append({
            "idx": i,
            "passo_indice": passo_indice,
            "passo_titolo": passo_titolo,
            "da": da,
            "a": a,
            "motivo": motivo,
        })

    # Stato iniziale: organismo pre-esistente (continuità) SOLO per riconferme
    # e già attivati; le nuove iscrizioni non hanno ancora un organismo, così
    # la loro prima assegnazione non viene contata come uno spostamento.
    stato_iniziale = {}
    for i in df_work.index:
        if is_fisso.at[i]:
            v = df_work.at[i, "_org_orig"]
            stato_iniziale[i] = v.strip() if isinstance(v, str) and v.strip() else ""
        else:
            stato_iniziale[i] = ""
    _registra_passo(
        "Stato iniziale (istanza MESIS)",
        "Organismo risultante dal MESIS all'avvio del procedimento: per le "
        "riconferme e gli alunni già attivati è l'organismo pre-esistente "
        "(continuità educativa); le nuove iscrizioni non hanno ancora un "
        "organismo assegnato.",
        "Dati di input — estrazione MESIS",
        stato=stato_iniziale,
    )

    for idx in df_work.index[is_altro]:
        tipo_orig = df_work.at[idx, tipo_col] if tipo_col else "?"
        df_work.at[idx, "_status"] = "Da assegnare manualmente"
        df_work.at[idx, "_note"] = f"Tipo iscrizione non riconosciuto: '{tipo_orig}'"
        df_work.at[idx, "_pref_soddisfatta"] = "Non assegnato"

    for idx in df_work.index[is_riconferma]:
        org = df_work.at[idx, "_org_orig"]
        if isinstance(org, str) and org.strip():
            df_work.at[idx, "_assegnato"] = org.strip()
            df_work.at[idx, "_pref_soddisfatta"] = "Riconferma"
            df_work.at[idx, "_status"] = "OK"
        else:
            df_work.at[idx, "_status"] = "OK"
            df_work.at[idx, "_pref_soddisfatta"] = "Riconferma"
            df_work.at[idx, "_note"] = "Riconferma senza organismo pre-esistente"

    for idx in df_work.index[is_attivato]:
        df_work.at[idx, "_assegnato"] = df_work.at[idx, "_org_orig"].strip()
        df_work.at[idx, "_pref_soddisfatta"] = "Già attivato"
        df_work.at[idx, "_status"] = "OK"

    nuove_idx = sorted(
        df_work.index[is_nuova_da_assegnare].tolist(),
        key=lambda i: df_work.at[i, "_codice"],
    )

    for idx in nuove_idx:
        prefs = df_work.at[idx, "_preferenze"]
        if not prefs:
            df_work.at[idx, "_status"] = "Preferenze non espresse"
            df_work.at[idx, "_note"] = "Preferenze non espresse — richiedere alla famiglia"
            df_work.at[idx, "_pref_soddisfatta"] = "Non assegnato"
            continue
        df_work.at[idx, "_assegnato"] = prefs[0]
        df_work.at[idx, "_pref_idx"] = 0

    _registra_passo(
        "Assegnazione iniziale",
        "Riconferme e alunni già attivati confermati sul proprio organismo "
        "per continuità educativa; nuove iscrizioni assegnate alla 1ª "
        "preferenza espressa dalla famiglia.",
        "Art. 5, comma 6 (continuità educativa), Linee Guida DGC Roma "
        "Capitale n. 260/2024",
    )

    gruppi = df_work["_gruppo"].unique()
    ore_totali_gruppo: dict[str, float] = {}
    for g in gruppi:
        ore_totali_gruppo[g] = df_work.loc[df_work["_gruppo"] == g, "_ore"].sum()

    # Vincolo 45h (Art. 5, comma 5): una cooperativa è ammissibile in un
    # gruppo solo se il totale delle ore ad essa assegnate — riconferme e
    # alunni già attivati COMPRESI (le loro ore concorrono alla soglia) —
    # raggiunge la soglia. La sola presenza di una riconferma NON rende
    # ammissibile la coop se il totale resta sotto soglia. Ogni nuova
    # iscrizione la cui cooperativa non raggiunge la soglia scorre in ordine
    # alla preferenza successiva; il puntatore di preferenza (_pref_idx)
    # avanza sempre in avanti, così la procedura converge in un numero finito
    # di passi e la destinazione finale raggiunge effettivamente la soglia
    # (o, esaurite le preferenze, la domanda va assegnata manualmente).
    for iteration in range(1, max_iter + 1):
        spostamenti = 0

        ore_per_coop: dict[str, dict[str, float]] = {g: {} for g in gruppi}
        for idx in df_work.index:
            if df_work.at[idx, "_status"] in (
                "Preferenze non espresse",
                "Da assegnare manualmente",
            ):
                continue
            g = df_work.at[idx, "_gruppo"]
            org = df_work.at[idx, "_assegnato"]
            if not org:
                continue
            org_n = normalizza_nome(org)
            ore_per_coop[g][org_n] = ore_per_coop[g].get(org_n, 0) + df_work.at[idx, "_ore"]

        passo_idx_iter = len(cronologia_passi)
        titolo_iter = f"Applicazione vincolo {soglia_ore}h — iterazione {iteration}"

        for idx in nuove_idx:
            if df_work.at[idx, "_status"] in (
                "Preferenze non espresse",
                "Da assegnare manualmente",
            ):
                continue
            g = df_work.at[idx, "_gruppo"]
            # eccezione: nei gruppi con monte ore totale sotto soglia la soglia
            # non è raggiungibile da nessuna coop — l'assegnazione resta
            if ore_totali_gruppo[g] < soglia_ore:
                continue
            org = df_work.at[idx, "_assegnato"]
            if not org:
                continue
            org_n = normalizza_nome(org)
            if ore_per_coop[g].get(org_n, 0) >= soglia_ore:
                continue  # la cooperativa attuale raggiunge le 45h: confermata

            # cooperativa attuale sotto soglia (totale, riconferme comprese):
            # scorri alla preferenza successiva espressa dalla famiglia
            g_label = cronologia_anagrafica[idx]["gruppo_label"]
            prefs = df_work.at[idx, "_preferenze"]
            pi = df_work.at[idx, "_pref_idx"] + 1
            while pi < len(prefs) and normalizza_nome(prefs[pi]) == org_n:
                pi += 1  # salta eventuali ripetizioni della stessa coop
            if pi < len(prefs):
                nuovo = prefs[pi]
                df_work.at[idx, "_assegnato"] = nuovo
                df_work.at[idx, "_pref_idx"] = pi
                _registra_movimento(
                    idx, passo_idx_iter, titolo_iter, org, nuovo,
                    f"L'organismo «{org}» non raggiunge la soglia di "
                    f"{soglia_ore}h nel gruppo «{g_label}» (somma di tutte le "
                    "ore assegnate, riconferme comprese): la domanda è "
                    f"riassegnata alla {pi + 1}ª preferenza espressa.",
                )
                spostamenti += 1
            else:
                df_work.at[idx, "_assegnato"] = ""
                df_work.at[idx, "_status"] = "Da assegnare manualmente"
                df_work.at[idx, "_note"] = (
                    "Tutte le preferenze espresse risultano sotto soglia "
                    f"{soglia_ore}h nel gruppo — assegnazione manuale necessaria"
                )
                df_work.at[idx, "_pref_soddisfatta"] = "Non assegnato"
                _registra_movimento(
                    idx, passo_idx_iter, titolo_iter, org, "",
                    f"Tutte le preferenze espresse risultano sotto la soglia "
                    f"di {soglia_ore}h nel gruppo «{g_label}»: la domanda "
                    "richiede assegnazione manuale.",
                )
                spostamenti += 1

        stats["iterazioni"] = iteration
        if spostamenti == 0:
            log.append(f"Iterazione {iteration}: nessuno spostamento — convergenza raggiunta.")
            break

        log.append(
            f"Iterazione {iteration}: {spostamenti} spostamenti, "
            f"{sum(1 for i in nuove_idx if df_work.at[i, '_status'] == 'Da assegnare manualmente')} "
            f"casi non risolti."
        )
        _registra_passo(
            titolo_iter,
            f"{spostamenti} spostamenti: le nuove iscrizioni la cui "
            f"cooperativa non raggiunge {soglia_ore} ore settimanali nel "
            "gruppo (contando tutte le ore assegnate, riconferme comprese) "
            "sono state riassegnate alla preferenza successiva espressa; le "
            "riconferme e gli alunni già attivati non vengono mai spostati.",
            f"Art. 5, comma 5, Linee Guida DGC Roma Capitale n. 260/2024 "
            f"(soglia minima {soglia_ore}h per organismo)",
        )
        stats["spostamenti"] += spostamenti
    else:
        log.append(f"Raggiunto limite massimo di {max_iter} iterazioni.")

    for idx in nuove_idx:
        if df_work.at[idx, "_status"] in ("Preferenze non espresse", "Da assegnare manualmente"):
            continue
        prefs = df_work.at[idx, "_preferenze"]
        pi = df_work.at[idx, "_pref_idx"]
        df_work.at[idx, "_pref_soddisfatta"] = f"{pi + 1}ª"
        df_work.at[idx, "_status"] = "OK"

    ore_per_coop_final: dict[str, dict[str, float]] = {g: {} for g in gruppi}
    for idx in df_work.index:
        if df_work.at[idx, "_status"] in ("Preferenze non espresse", "Da assegnare manualmente"):
            continue
        g = df_work.at[idx, "_gruppo"]
        org = df_work.at[idx, "_assegnato"]
        if not org:
            continue
        org_n = normalizza_nome(org)
        ore_per_coop_final[g][org_n] = (
            ore_per_coop_final[g].get(org_n, 0) + df_work.at[idx, "_ore"]
        )

    for idx in df_work.index[is_fisso]:
        g = df_work.at[idx, "_gruppo"]
        org = df_work.at[idx, "_assegnato"]
        if not org:
            continue
        org_n = normalizza_nome(org)
        soglia_attiva = ore_totali_gruppo.get(g, 0) >= soglia_ore
        ore_coop = ore_per_coop_final.get(g, {}).get(org_n, 0)
        if soglia_attiva and ore_coop < soglia_ore:
            if is_riconferma.at[idx]:
                df_work.at[idx, "_status"] = "Riconferma sotto soglia"
                df_work.at[idx, "_note"] = (
                    f"Riconferma sotto soglia {soglia_ore}h — "
                    "verificare con la Direzione Socio-Educativa municipale"
                )
            else:
                df_work.at[idx, "_status"] = "Attivato sotto soglia"
                df_work.at[idx, "_note"] = (
                    f"Alunno già attivato con organismo sotto soglia {soglia_ore}h — "
                    "verificare con la Direzione Socio-Educativa municipale"
                )

    stats["criticita"] = int(
        (df_work["_status"] != "OK").sum()
    )

    _registra_passo(
        "Esito definitivo",
        "Assegnazione definitiva risultante dal procedimento automatizzato, "
        "dopo la convergenza del vincolo delle 45 ore settimanali per "
        "organismo. Eventuali casi non risolti sono elencati tra le criticità.",
        "Esito del procedimento automatizzato",
    )

    stats["cronologia"] = {
        "passi": cronologia_passi,
        "movimenti": cronologia_movimenti,
        "anagrafica": cronologia_anagrafica,
        "soglia_ore": soglia_ore,
        "corso_anno": corso_anno,
    }

    dn_col = col("data_nascita")
    plesso_nome_col = col("plesso")
    istituto_col = col("istituto")
    codice_mecc_ist_col = col("codice_mecc_istituto")
    classe_col = col("classe")
    sezione_col = col("sezione")
    ambito_col = col("ambito")
    ore_rich_col = col("ore_richieste")
    data_att_col = col("data_attivazione")
    municipio_col = col("municipio")

    result_rows = []
    for idx in df_work.index:
        dn = parse_data_nascita(df_work.at[idx, dn_col]) if dn_col else None
        eta = calcola_eta(dn)
        plesso_code = df_work.at[idx, "_plesso"]
        ambito_val = df_work.at[idx, "_ambito"]
        tipo_gest = df_work.at[idx, "_tipo_gest"]
        grado = deriva_grado(plesso_code, tipo_gest if tipo_gest != "N/D" else ambito_val)

        dn_str = dn.strftime("%d/%m/%Y") if dn else ""

        result_rows.append({
            "Codice Iscrizione": df_work.at[idx, "_codice"],
            "Tipo": df_work.at[idx, tipo_col] if tipo_col else "",
            "Cognome": df_work.at[idx, col("cognome")] if col("cognome") else "",
            "Nome": df_work.at[idx, col("nome")] if col("nome") else "",
            "Data Nascita": dn_str,
            "Età": eta if eta is not None else "",
            "Codice Meccanografico Istituto": (
                df_work.at[idx, codice_mecc_ist_col] if codice_mecc_ist_col else ""
            ),
            "Istituto": df_work.at[idx, istituto_col] if istituto_col else "",
            "Codice Meccanografico Plesso": plesso_code,
            "Plesso": df_work.at[idx, plesso_nome_col] if plesso_nome_col else "",
            "Indirizzo Plesso": "",
            "Grado Scolastico": grado,
            "Classe": df_work.at[idx, classe_col] if classe_col else "",
            "Sezione": df_work.at[idx, sezione_col] if sezione_col else "",
            "Ambito": df_work.at[idx, "_ambito"],
            "Tipo Gestione": tipo_gest,
            "Fonte Classificazione": df_work.at[idx, "_fonte_class"],
            "Municipio": df_work.at[idx, "_municipio"],
            "Gruppo 45h": df_work.at[idx, "_gruppo"],
            "Ore Richieste": df_work.at[idx, ore_rich_col] if ore_rich_col else "",
            "Ore Assegnate": df_work.at[idx, "_ore"],
            "Organismo Pre-esistente": df_work.at[idx, "_org_orig"] if isinstance(df_work.at[idx, "_org_orig"], str) else "",
            "Organismo Assegnato dall'Algoritmo": df_work.at[idx, "_assegnato"],
            "Preferenza Soddisfatta": df_work.at[idx, "_pref_soddisfatta"],
            "Data Attivazione": df_work.at[idx, data_att_col] if data_att_col else "",
            "Status": df_work.at[idx, "_status"],
            "Note": df_work.at[idx, "_note"],
        })

    df_result = pd.DataFrame(result_rows)

    sort_cols = [
        c for c in [
            "Codice Meccanografico Istituto",
            "Codice Meccanografico Plesso",
            "Classe",
            "Sezione",
            "Cognome",
            "Nome",
        ]
        if c in df_result.columns
    ]
    if sort_cols:
        if "Classe" in df_result.columns:
            df_result["_classe_sort"] = pd.to_numeric(
                df_result["Classe"], errors="coerce"
            ).fillna(float("inf"))
            real_sort = [
                "_classe_sort" if c == "Classe" else c for c in sort_cols
            ]
            df_result.sort_values(real_sort, inplace=True, ignore_index=True)
            df_result.drop(columns=["_classe_sort"], inplace=True)
        else:
            df_result.sort_values(sort_cols, inplace=True, ignore_index=True)

    df_riepilogo = costruisci_riepilogo_gruppo(df_result, soglia_ore)
    df_criticita = costruisci_criticita(df_result)

    return df_result, df_riepilogo, df_criticita, log, stats, nome_registro


def costruisci_riepilogo_gruppo(df_result, soglia_ore=45):
    """Costruisce il riepilogo per (Gruppo 45h x Organismo) dal DataFrame assegnazioni."""
    ORG = "Organismo Assegnato dall'Algoritmo"

    def _nomi_scuole(df_g, colonna):
        nomi = sorted(
            n for n in df_g[colonna].fillna("").astype(str).str.strip().unique() if n
        ) if colonna in df_g.columns else []
        return ", ".join(nomi[:3]) + (f" (+{len(nomi)-3})" if len(nomi) > 3 else "")

    riepilogo_rows = []
    gruppi_result = df_result["Gruppo 45h"].unique()
    for grp in sorted(g for g in gruppi_result if g):
        df_g = df_result.loc[df_result["Gruppo 45h"] == grp]
        ore_tot_g = df_g["Ore Assegnate"].sum()
        n_plessi = df_g["Codice Meccanografico Plesso"].nunique()

        if grp.startswith("IC:"):
            tipo_gruppo = "IC"
            desc_gruppo = df_g["Istituto"].iloc[0] if "Istituto" in df_g.columns else ""
        elif grp.startswith("COM:"):
            tipo_gruppo = "Infanzia comunale"
            desc_gruppo = _nomi_scuole(df_g, "Plesso")
        elif grp.startswith("PAR:"):
            tipo_gruppo = "Paritario"
            desc_gruppo = _nomi_scuole(df_g, "Plesso")
        elif grp.startswith("AMB:"):
            amb_val = grp.replace("AMB:", "")
            tipo_gruppo = "Paritario" if "Paritario" in amb_val else "Comunale"
            nomi = sorted(df_g.drop_duplicates("Codice Meccanografico Istituto")["Istituto"].unique())
            desc_gruppo = ", ".join(nomi[:3]) + (f" (+{len(nomi)-3})" if len(nomi) > 3 else "")
        else:
            tipo_gruppo = ""
            desc_gruppo = df_g["Plesso"].iloc[0] if "Plesso" in df_g.columns else ""
            n_plessi = 1

        if "Municipio" in df_g.columns:
            municipi = sorted(m for m in df_g["Municipio"].unique() if m)
            municipio_str = ", ".join(municipi)
        else:
            municipio_str = ""

        for org in sorted(o for o in df_g[ORG].unique() if o):
            mask_o = df_g[ORG] == org
            ore_org = df_g.loc[mask_o, "Ore Assegnate"].sum()
            if ore_tot_g < soglia_ore:
                soglia_str = "N/A (gruppo sotto 45h totali)"
            elif ore_org >= soglia_ore:
                soglia_str = "Raggiunta"
            else:
                soglia_str = "Non raggiunta"
            riepilogo_rows.append({
                "Gruppo 45h": grp,
                "Tipo gruppo": tipo_gruppo,
                "Descrizione": desc_gruppo,
                "Ambito": df_g["Ambito"].iloc[0],
                "Municipio": municipio_str,
                "N. plessi": n_plessi,
                "Organismo": org,
                "N. alunni assegnati": int(mask_o.sum()),
                "Ore totali settimanali org.": ore_org,
                "Ore totali del gruppo": ore_tot_g,
                "Soglia 45h": soglia_str,
            })
    return pd.DataFrame(riepilogo_rows)


def costruisci_criticita(df_result):
    """Estrae i record critici (Status != OK) con azione suggerita."""
    df_criticita = df_result.loc[df_result["Status"] != "OK"].copy()
    azioni = {
        "Preferenze non espresse": "Contattare la famiglia per acquisire le preferenze",
        "Da assegnare manualmente": "Spostamento manuale necessario",
        "Riconferma sotto soglia": "Inoltrare segnalazione formale alla Direzione Socio-Educativa",
        "Attivato sotto soglia": "Inoltrare segnalazione formale alla Direzione Socio-Educativa",
    }
    if not df_criticita.empty:
        df_criticita["Azione suggerita"] = df_criticita["Status"].map(azioni).fillna("")
    else:
        df_criticita["Azione suggerita"] = []
    return df_criticita


def costruisci_tabella_scuole(
    df_raw: pd.DataFrame,
    col_map: dict[str, str],
    codici_comunali: set[str] | None = None,
) -> pd.DataFrame:
    """Tabella dei plessi (iscrizioni attive) con la classificazione derivata.

    Serve per la verifica/correzione manuale prima dell'assegnazione:
    una riga per codice meccanografico plesso.
    """
    def col(key):
        return col_map.get(key)

    stato_col = col("stato")
    plesso_col = col("codice_mecc_plesso")
    if not stato_col or not plesso_col:
        return pd.DataFrame()

    mask = df_raw[stato_col].str.strip().str.upper() == "ATTIVA"
    rinuncia_col = col("data_rinuncia")
    if rinuncia_col:
        mask &= df_raw[rinuncia_col].isna() | (df_raw[rinuncia_col].str.strip() == "")
    df_a = df_raw.loc[mask].copy()
    if df_a.empty:
        return pd.DataFrame()

    ore_col = col("ore_assegnate")
    df_a["_plesso"] = df_a[plesso_col].str.strip().fillna("")
    df_a["_ore"] = (
        pd.to_numeric(df_a[ore_col], errors="coerce").fillna(0) if ore_col else 0
    )

    # stessa risoluzione per plesso usata dal motore di assegnazione:
    # la tabella mostra esattamente i gruppi che verranno applicati
    risoluzione = risolvi_plessi(
        df_a,
        col_plesso="_plesso",
        col_ambito=col("ambito"),
        col_istituto_nome=col("istituto"),
        col_plesso_nome=col("plesso"),
        col_municipio=col("municipio"),
        col_istituto_cod=col("codice_mecc_istituto"),
        codici_comunali=codici_comunali,
    )

    rows = []
    for cod, df_p in df_a.groupby("_plesso", sort=True):
        info = risoluzione.get(str(cod).strip())
        if info is None:
            continue
        rows.append({
            "Codice Meccanografico Plesso": str(cod).strip(),
            "Plesso": info["plesso"],
            "Istituto": info["istituto"],
            "Municipio": info["municipio"],
            "Ambito MESIS": info["ambito"],
            "Tipo Gestione": info["tipo"],
            "Fonte": info["fonte"],
            "Gruppo 45h (auto)": info["gruppo"],
            "N. alunni": len(df_p),
            "Ore": df_p["_ore"].sum(),
        })
    return pd.DataFrame(rows)


def arricchisci_indirizzi(df_result: pd.DataFrame, file_bytes: bytes, filename: str = "") -> pd.DataFrame:
    try:
        if filename.lower().endswith(".csv"):
            df_ana = pd.read_csv(io.BytesIO(file_bytes), dtype=str)
        else:
            try:
                df_ana = pd.read_excel(io.BytesIO(file_bytes), dtype=str)
            except Exception:
                df_ana = pd.read_csv(io.BytesIO(file_bytes), dtype=str)
        cols_lower = {c.strip().lower(): c for c in df_ana.columns}

        cod_col = None
        ind_col = None
        for k, v in cols_lower.items():
            if "codice" in k and "meccanografico" in k:
                cod_col = v
            elif "codice_meccanografico" in k:
                cod_col = v
            elif k == "codicescuola":
                cod_col = v
        for k, v in cols_lower.items():
            if "indirizzo" in k:
                ind_col = v

        if cod_col and ind_col:
            lookup = dict(
                zip(
                    df_ana[cod_col].str.strip(),
                    df_ana[ind_col].str.strip(),
                )
            )
            df_result["Indirizzo Plesso"] = df_result[
                "Codice Meccanografico Plesso"
            ].map(lookup).fillna("")
    except Exception:
        pass
    return df_result


# ---------------------------------------------------------------------------
# Economic calculations (shared by Excel, PDF, dashboard)
# ---------------------------------------------------------------------------

def calcola_colonne_economiche(
    df, n_settimane=35, perc_decurtazione=11.0, costo_orario=24.07, aliquota_iva=5.0
):
    """Aggiunge le colonne economiche derivate al DataFrame delle assegnazioni."""
    coeff_netto = 1 - perc_decurtazione / 100
    coeff_iva = aliquota_iva / 100
    ore_sett = pd.to_numeric(df["Ore Assegnate"], errors="coerce").fillna(0)
    ore_annuali = ore_sett * n_settimane
    decurtazione = ore_annuali * (perc_decurtazione / 100)
    ore_nette = ore_annuali * coeff_netto
    imponibile = ore_nette * costo_orario
    iva = imponibile * coeff_iva
    totale = imponibile + iva
    out = df.copy()
    out["Ore annuali lorde"] = ore_annuali
    out[f"Decurtazione {perc_decurtazione:.0f}%"] = decurtazione
    out["Ore annuali nette"] = ore_nette
    out["Imponibile (EUR)"] = imponibile.round(2)
    out[f"IVA {aliquota_iva:.0f}% (EUR)"] = iva.round(2)
    out["Totale (EUR)"] = totale.round(2)
    return out


def calcola_riepilogo_economico(
    df_eco, n_settimane=35, perc_decurtazione=11.0, costo_orario=24.07, aliquota_iva=5.0
):
    """Costruisce il riepilogo economico aggregato per organismo."""
    ORG = "Organismo Assegnato dall'Algoritmo"
    organismi = sorted([o for o in df_eco[ORG].unique() if o])
    iva_col = f"IVA {aliquota_iva:.0f}% (EUR)"
    rows = []
    for org in organismi:
        df_o = df_eco.loc[df_eco[ORG] == org]
        ore_lorde = df_o["Ore annuali lorde"].sum()
        ore_nette = df_o["Ore annuali nette"].sum()
        rows.append({
            "Organismo": org,
            "N. alunni": len(df_o),
            "Ore sett. totali": df_o["Ore Assegnate"].sum(),
            f"Ore annuali lorde ({n_settimane} sett.)": ore_lorde,
            f"Decurtazione {perc_decurtazione:.0f}%": ore_lorde - ore_nette,
            "Ore annuali nette": ore_nette,
            "Costo orario": costo_orario,
            "Imponibile (EUR)": round(df_o["Imponibile (EUR)"].sum(), 2),
            iva_col: round(df_o[iva_col].sum(), 2),
            "Totale (EUR)": round(df_o["Totale (EUR)"].sum(), 2),
        })
    return pd.DataFrame(rows)


def colonne_coop(perc_decurtazione=11.0, aliquota_iva=5.0):
    """Colonne da mostrare nei fogli/lettere per cooperativa."""
    return [
        "Codice Iscrizione", "Tipo", "Cognome", "Nome", "Data Nascita", "Età",
        "Codice Meccanografico Istituto", "Istituto",
        "Codice Meccanografico Plesso", "Plesso", "Grado Scolastico",
        "Classe", "Sezione", "Ambito", "Tipo Gestione", "Municipio", "Gruppo 45h",
        "Ore Assegnate", "Ore annuali lorde",
        f"Decurtazione {perc_decurtazione:.0f}%", "Ore annuali nette",
        "Imponibile (EUR)", f"IVA {aliquota_iva:.0f}% (EUR)", "Totale (EUR)",
        "Preferenza Soddisfatta", "Status",
    ]


# ---------------------------------------------------------------------------
# Excel output with formatting
# ---------------------------------------------------------------------------

def genera_excel(
    df_assegnazioni, df_riepilogo, df_criticita,
    n_settimane: int = 35,
    perc_decurtazione: float = 11.0,
    costo_orario: float = 24.07,
    aliquota_iva: float = 5.0,
) -> bytes:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    BRAND = "1F4E79"
    BRAND_LIGHT = "D6E4F0"
    ACCENT = "2E75B6"
    WHITE = "FFFFFF"
    ZEBRA = "F2F7FB"

    HDR_FONT = Font(bold=True, color=WHITE, size=10)
    HDR_FILL = PatternFill(start_color=BRAND, end_color=BRAND, fill_type="solid")
    TITLE_FONT = Font(bold=True, size=14, color=BRAND)
    SUBTITLE_FONT = Font(bold=True, size=11, color=ACCENT)
    PARAM_FONT = Font(size=10, color="444444")
    TOTAL_FONT = Font(bold=True, size=10, color=BRAND)
    TOTAL_FILL = PatternFill(start_color=BRAND_LIGHT, end_color=BRAND_LIGHT, fill_type="solid")
    ZEBRA_FILL = PatternFill(start_color=ZEBRA, end_color=ZEBRA, fill_type="solid")
    STATUS_FILLS = {
        "OK": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
        "Riconferma sotto soglia": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
        "Attivato sotto soglia": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
        "Preferenze non espresse": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
        "Da assegnare manualmente": PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid"),
    }
    SOGLIA_FILLS = {
        "Raggiunta": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
        "Non raggiunta": PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid"),
    }
    THIN = Border(
        left=Side(style="thin", color="B4C6E7"),
        right=Side(style="thin", color="B4C6E7"),
        top=Side(style="thin", color="B4C6E7"),
        bottom=Side(style="thin", color="B4C6E7"),
    )
    EUR_FMT = '#,##0.00\ "EUR"'

    def _fmt_sheet(ws, n_cols, fill_map=None, fill_col_idx=None, start_row=1):
        for ci in range(1, n_cols + 1):
            cell = ws.cell(row=start_row, column=ci)
            cell.font = HDR_FONT
            cell.fill = HDR_FILL
            cell.border = THIN
            cell.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
        ws.row_dimensions[start_row].height = 28

        for ri in range(start_row + 1, ws.max_row + 1):
            is_zebra = (ri - start_row) % 2 == 0
            status_fill = None
            if fill_col_idx and fill_map:
                val = ws.cell(row=ri, column=fill_col_idx).value
                status_fill = fill_map.get(val)
            for ci in range(1, n_cols + 1):
                cell = ws.cell(row=ri, column=ci)
                cell.border = THIN
                if status_fill:
                    cell.fill = status_fill
                elif is_zebra:
                    cell.fill = ZEBRA_FILL

        for ci in range(1, n_cols + 1):
            max_len = 0
            for ri in range(start_row, min(ws.max_row + 1, 300)):
                val = ws.cell(row=ri, column=ci).value
                if val is not None:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 4, 45)
        ws.freeze_panes = ws.cell(row=start_row + 1, column=1).coordinate

    def _fmt_eur_cols(ws, col_names, start_row=1):
        for ci, name in enumerate(col_names, 1):
            if any(k in str(name) for k in ("EUR", "Imponibile", "Totale", "IVA")):
                for ri in range(start_row + 1, ws.max_row + 1):
                    ws.cell(row=ri, column=ci).number_format = EUR_FMT

    def _add_total_row(ws, col_names, data_start_row, sum_cols_keywords):
        sr = ws.max_row + 1
        ws.cell(row=sr, column=1).value = "TOTALE"
        ws.cell(row=sr, column=1).font = TOTAL_FONT
        for ci, name in enumerate(col_names, 1):
            if any(k in str(name) for k in sum_cols_keywords):
                total = sum(ws.cell(row=r, column=ci).value or 0 for r in range(data_start_row, sr))
                cell = ws.cell(row=sr, column=ci)
                cell.value = round(total, 2) if "EUR" in str(name) else total
                cell.font = TOTAL_FONT
                if "EUR" in str(name):
                    cell.number_format = EUR_FMT
            ws.cell(row=sr, column=ci).fill = TOTAL_FILL
            ws.cell(row=sr, column=ci).border = THIN
        return sr

    df_eco = calcola_colonne_economiche(
        df_assegnazioni, n_settimane, perc_decurtazione, costo_orario, aliquota_iva
    )

    organismi = sorted([o for o in df_eco["Organismo Assegnato dall'Algoritmo"].unique() if o])

    df_economico = calcola_riepilogo_economico(
        df_eco, n_settimane, perc_decurtazione, costo_orario, aliquota_iva
    )

    coop_cols = colonne_coop(perc_decurtazione, aliquota_iva)
    coop_cols_available = [c for c in coop_cols if c in df_eco.columns]

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_eco.to_excel(writer, sheet_name="Assegnazioni", index=False)
        df_riepilogo.to_excel(writer, sheet_name="Riepilogo Gruppo", index=False)
        df_economico.to_excel(writer, sheet_name="Riepilogo Economico", index=False)
        df_criticita.to_excel(writer, sheet_name="Criticita", index=False)

        for org in organismi:
            mask = df_eco["Organismo Assegnato dall'Algoritmo"] == org
            df_o = df_eco.loc[mask, coop_cols_available].copy()

            safe_name = re.sub(r"[\\/*?\[\]:]", "", org)[:28].strip()
            if not safe_name:
                safe_name = "Organismo"
            seen_names = [s.title for s in writer.sheets.values()] if hasattr(writer, 'sheets') else []
            final_name = safe_name
            counter = 2
            while final_name in seen_names:
                final_name = f"{safe_name[:25]}_{counter}"
                counter += 1

            df_o.to_excel(writer, sheet_name=final_name, index=False, startrow=5)

    output.seek(0)
    wb = load_workbook(output)

    ws_ass = wb["Assegnazioni"]
    status_idx = list(df_eco.columns).index("Status") + 1 if "Status" in df_eco.columns else None
    _fmt_sheet(ws_ass, ws_ass.max_column, STATUS_FILLS, status_idx)
    _fmt_eur_cols(ws_ass, df_eco.columns)

    ws_grp = wb["Riepilogo Gruppo"]
    soglia_idx = list(df_riepilogo.columns).index("Soglia 45h") + 1 if "Soglia 45h" in df_riepilogo.columns else None
    _fmt_sheet(ws_grp, ws_grp.max_column, SOGLIA_FILLS, soglia_idx)

    ws_eco = wb["Riepilogo Economico"]
    _fmt_sheet(ws_eco, ws_eco.max_column)
    _fmt_eur_cols(ws_eco, df_economico.columns)
    _add_total_row(ws_eco, list(df_economico.columns), 2,
                   ("alunni", "Ore", "EUR", "IVA", "Imponibile", "Totale", "Decurtazione"))

    ws_crit = wb["Criticita"]
    status_idx_c = list(df_criticita.columns).index("Status") + 1 if "Status" in df_criticita.columns and len(df_criticita) > 0 else None
    _fmt_sheet(ws_crit, ws_crit.max_column, STATUS_FILLS, status_idx_c)

    today_str = datetime.date.today().strftime("%d/%m/%Y")
    for org in organismi:
        safe_name = re.sub(r"[\\/*?\[\]:]", "", org)[:28].strip() or "Organismo"
        ws = None
        for s in wb.sheetnames:
            if s.startswith(safe_name[:20]):
                ws = wb[s]
                break
        if ws is None:
            continue

        mask = df_eco["Organismo Assegnato dall'Algoritmo"] == org
        df_o = df_eco.loc[mask]
        ore_sett_tot = df_o["Ore Assegnate"].sum()
        n_al = len(df_o)
        imponibile_tot = df_o["Imponibile (EUR)"].sum()
        iva_tot = df_o[f"IVA {aliquota_iva:.0f}% (EUR)"].sum()
        totale_tot = df_o["Totale (EUR)"].sum()

        ore_lorde_tot = df_o["Ore annuali lorde"].sum()
        ore_nette_tot = df_o["Ore annuali nette"].sum()

        ws.cell(row=1, column=1).value = "ASSEGNAZIONE SERVIZIO OEPAC"
        ws.cell(row=1, column=1).font = TITLE_FONT
        ws.cell(row=2, column=1).value = org
        ws.cell(row=2, column=1).font = SUBTITLE_FONT
        ws.cell(row=3, column=1).value = (
            f"Data: {today_str}   |   Alunni assegnati: {n_al}   |   "
            f"Ore settimanali: {ore_sett_tot:.0f}   |   "
            f"Settimane: {n_settimane}   |   "
            f"Ore annuali nette (decurt. {perc_decurtazione:.0f}%): {ore_nette_tot:,.0f}"
        )
        ws.cell(row=3, column=1).font = PARAM_FONT
        ws.cell(row=4, column=1).value = (
            f"Costo orario: EUR {costo_orario:.2f}   |   "
            f"Imponibile: EUR {imponibile_tot:,.2f}   |   "
            f"IVA {aliquota_iva:.0f}%: EUR {iva_tot:,.2f}   |   "
            f"TOTALE: EUR {totale_tot:,.2f}"
        )
        ws.cell(row=4, column=1).font = SUBTITLE_FONT

        n_data_cols = len(coop_cols_available)
        _fmt_sheet(ws, n_data_cols, None, None, start_row=6)
        _fmt_eur_cols(ws, coop_cols_available, start_row=6)
        _add_total_row(ws, coop_cols_available, 7,
                       ("Ore", "EUR", "IVA", "Imponibile", "Totale", "Decurtazione"))

    final = io.BytesIO()
    wb.save(final)
    final.seek(0)
    return final.getvalue()


def _safe_filename(nome: str) -> str:
    """Nome file sicuro derivato dal nome organismo."""
    s = re.sub(r"[\\/*?\[\]:<>|\"']", "", nome).strip()
    s = re.sub(r"\s+", "_", s)
    return s[:60] or "Organismo"


def genera_zip_excel_cooperative(
    df_assegnazioni, df_riepilogo, df_criticita,
    n_settimane=35, perc_decurtazione=11.0, costo_orario=24.07, aliquota_iva=5.0,
) -> bytes:
    """Genera uno ZIP con un file .xlsx separato per ogni cooperativa
    (contenente SOLO i suoi alunni — privacy GDPR)."""
    import zipfile

    ORG = "Organismo Assegnato dall'Algoritmo"
    organismi = sorted([o for o in df_assegnazioni[ORG].unique() if o])

    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for org in organismi:
            mask = df_assegnazioni[ORG] == org
            df_ass_o = df_assegnazioni.loc[mask].copy()
            gruppi_o = set(df_ass_o["Gruppo 45h"].unique())
            df_riep_o = df_riepilogo.loc[
                (df_riepilogo["Organismo"] == org)
                & (df_riepilogo["Gruppo 45h"].isin(gruppi_o))
            ].copy()
            df_crit_o = df_criticita.loc[
                df_criticita[ORG] == org
            ].copy() if ORG in df_criticita.columns else df_criticita.iloc[0:0].copy()

            xlsx_bytes = genera_excel(
                df_ass_o, df_riep_o, df_crit_o,
                n_settimane, perc_decurtazione, costo_orario, aliquota_iva,
            )
            zf.writestr(f"OEPAC_{_safe_filename(org)}.xlsx", xlsx_bytes)

    zip_buf.seek(0)
    return zip_buf.getvalue()


def genera_pdf_coop(
    df_eco_org, org, n_settimane=35, perc_decurtazione=11.0,
    costo_orario=24.07, aliquota_iva=5.0,
) -> bytes:
    """Genera una lettera di assegnazione PDF per una cooperativa."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    )

    BRAND = colors.HexColor("#1F4E79")
    BRAND_LIGHT = colors.HexColor("#D6E4F0")
    ZEBRA = colors.HexColor("#F2F7FB")

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=12 * mm, rightMargin=12 * mm,
        topMargin=14 * mm, bottomMargin=12 * mm,
    )
    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("h1", parent=styles["Title"], fontSize=16,
                        textColor=BRAND, spaceAfter=2)
    h2 = ParagraphStyle("h2", parent=styles["Heading2"], fontSize=12,
                        textColor=colors.HexColor("#2E75B6"), spaceAfter=8)
    normal = ParagraphStyle("n", parent=styles["Normal"], fontSize=9, leading=13)
    small = ParagraphStyle("s", parent=styles["Normal"], fontSize=8,
                           textColor=colors.HexColor("#666666"))

    iva_col = f"IVA {aliquota_iva:.0f}% (EUR)"
    ore_sett = df_eco_org["Ore Assegnate"].sum()
    ore_nette = df_eco_org["Ore annuali nette"].sum()
    imponibile = df_eco_org["Imponibile (EUR)"].sum()
    iva = df_eco_org[iva_col].sum()
    totale = df_eco_org["Totale (EUR)"].sum()
    n_al = len(df_eco_org)
    today_str = datetime.date.today().strftime("%d/%m/%Y")

    elems = []
    elems.append(Paragraph("ASSEGNAZIONE SERVIZIO OEPAC", h1))
    elems.append(Paragraph(org, h2))
    elems.append(Paragraph(
        f"Roma, {today_str} &nbsp;&nbsp;|&nbsp;&nbsp; "
        "Ai sensi dell'Art. 5, commi 5 e 6 delle Linee Guida approvate con "
        "DGC Roma Capitale n. 260/2024.", normal,
    ))
    elems.append(Spacer(1, 6))
    elems.append(Paragraph(
        f"Si comunica l'assegnazione del servizio OEPAC per <b>{n_al} alunni/e</b>, "
        f"per un totale di <b>{ore_sett:.0f} ore settimanali</b>. "
        f"Il computo economico, calcolato su <b>{n_settimane} settimane</b> con "
        f"decurtazione del <b>{perc_decurtazione:.0f}%</b> e costo orario di "
        f"<b>EUR {costo_orario:.2f}</b>, e riportato di seguito.", normal,
    ))
    elems.append(Spacer(1, 8))

    riepilogo_data = [
        ["Ore sett.", "Ore annue nette", "Imponibile", f"IVA {aliquota_iva:.0f}%", "TOTALE"],
        [
            f"{ore_sett:.0f}",
            f"{ore_nette:,.0f}",
            f"EUR {imponibile:,.2f}",
            f"EUR {iva:,.2f}",
            f"EUR {totale:,.2f}",
        ],
    ]
    rt = Table(riepilogo_data, hAlign="LEFT")
    rt.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), BRAND),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, 1), (-1, 1), "Helvetica-Bold"),
        ("BACKGROUND", (0, 1), (-1, 1), BRAND_LIGHT),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#B4C6E7")),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    elems.append(rt)
    elems.append(Spacer(1, 12))
    elems.append(Paragraph("Dettaglio alunni assegnati", h2))

    cols = [
        "Cognome", "Nome", "Grado Scolastico", "Plesso", "Classe", "Sezione",
        "Ore Assegnate", "Ore annuali nette", "Imponibile (EUR)", "Totale (EUR)",
    ]
    cols = [c for c in cols if c in df_eco_org.columns]
    header = [c.replace(" (EUR)", "").replace("Grado Scolastico", "Grado")
              .replace("Ore Assegnate", "Ore/sett").replace("Ore annuali nette", "Ore/anno")
              for c in cols]

    df_sorted = df_eco_org.sort_values(
        [c for c in ["Plesso", "Cognome", "Nome"] if c in df_eco_org.columns]
    )
    data = [header]
    for _, r in df_sorted.iterrows():
        row = []
        for c in cols:
            v = r[c]
            if "EUR" in c:
                row.append(f"{v:,.2f}")
            elif c in ("Ore annuali nette",):
                row.append(f"{v:,.0f}")
            else:
                row.append(str(v) if pd.notna(v) else "")
        data.append(row)

    tot_row = ["TOTALE"] + [""] * (len(cols) - 1)
    for i, c in enumerate(cols):
        if c == "Ore Assegnate":
            tot_row[i] = f"{ore_sett:.0f}"
        elif c == "Ore annuali nette":
            tot_row[i] = f"{ore_nette:,.0f}"
        elif c == "Imponibile (EUR)":
            tot_row[i] = f"{imponibile:,.2f}"
        elif c == "Totale (EUR)":
            tot_row[i] = f"{totale:,.2f}"
    data.append(tot_row)

    n_cols = len(cols)
    page_w = landscape(A4)[0] - 24 * mm
    name_w = page_w * 0.13
    other_w = (page_w - 2 * name_w) / (n_cols - 2) if n_cols > 2 else page_w / n_cols
    col_widths = []
    for c in cols:
        col_widths.append(name_w if c in ("Cognome", "Nome", "Plesso") else other_w)

    t = Table(data, colWidths=col_widths, repeatRows=1)
    style = [
        ("BACKGROUND", (0, 0), (-1, 0), BRAND),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#B4C6E7")),
        ("ALIGN", (4, 1), (-1, -1), "CENTER"),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("BACKGROUND", (0, -1), (-1, -1), BRAND_LIGHT),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
    ]
    for ri in range(1, len(data) - 1):
        if ri % 2 == 0:
            style.append(("BACKGROUND", (0, ri), (-1, ri), ZEBRA))
    t.setStyle(TableStyle(style))
    elems.append(t)
    elems.append(Spacer(1, 10))
    elems.append(Paragraph(
        "Documento generato automaticamente dal sistema di assegnazione OEPAC. "
        "I dati riguardano alunni con disabilita e sono soggetti a riservatezza "
        "(Reg. UE 2016/679, art. 9).", small,
    ))

    doc.build(elems)
    buf.seek(0)
    return buf.getvalue()


def genera_zip_pdf_cooperative(
    df_assegnazioni, n_settimane=35, perc_decurtazione=11.0,
    costo_orario=24.07, aliquota_iva=5.0,
) -> bytes:
    """ZIP con una lettera PDF di assegnazione per ogni cooperativa."""
    import zipfile

    ORG = "Organismo Assegnato dall'Algoritmo"
    df_eco = calcola_colonne_economiche(
        df_assegnazioni, n_settimane, perc_decurtazione, costo_orario, aliquota_iva
    )
    organismi = sorted([o for o in df_eco[ORG].unique() if o])

    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for org in organismi:
            df_o = df_eco.loc[df_eco[ORG] == org].copy()
            pdf_bytes = genera_pdf_coop(
                df_o, org, n_settimane, perc_decurtazione, costo_orario, aliquota_iva
            )
            zf.writestr(f"Lettera_OEPAC_{_safe_filename(org)}.pdf", pdf_bytes)

    zip_buf.seek(0)
    return zip_buf.getvalue()


# ---------------------------------------------------------------------------
# Documentazione del procedimento (trasparenza / accesso agli atti)
# ---------------------------------------------------------------------------

PLACEHOLDER_NON_ASSEGNATO = "— (non assegnato)"
PLACEHOLDER_MANUALE = "— (assegnazione manuale)"


def costruisci_cronologia_dfs(cronologia: dict):
    """Dai dati grezzi della cronologia costruisce tre tabelle:

    - df_passi: una riga per fase del procedimento con conteggi;
    - df_diario: una riga per alunno con l'organismo ad ogni fase
      (il "prima e dopo" progressivo);
    - df_movimenti: una riga per ogni spostamento, con la motivazione.
    Ritorna anche l'elenco delle intestazioni delle colonne-fase del diario.
    """
    passi = cronologia.get("passi", [])
    movimenti = cronologia.get("movimenti", [])
    anagrafica = cronologia.get("anagrafica", {})

    def _chiave(idx):
        a = anagrafica[idx]
        return (
            str(a.get("gruppo_label", "")),
            str(a.get("cognome", "")).upper(),
            str(a.get("nome", "")).upper(),
        )

    ordine = sorted(anagrafica.keys(), key=_chiave)

    # --- Sintesi passaggi ---
    # Uno "spostamento" è una riassegnazione: l'alunno aveva già un organismo
    # ed è cambiato (o è tornato non assegnato). La prima assegnazione da vuoto
    # non è uno spostamento. Così la somma coincide con il registro movimenti.
    righe_passi = []
    for k, p in enumerate(passi):
        stato = p["stato"]
        n_ass = sum(1 for v in stato.values() if v)
        if k == 0:
            n_spost = ""
        else:
            prev = passi[k - 1]["stato"]
            n_spost = sum(
                1 for i in stato
                if prev.get(i, "") and stato.get(i, "") != prev.get(i, "")
            )
        righe_passi.append({
            "N.": p["indice"],
            "Fase": p["titolo"],
            "Descrizione": p["descrizione"],
            "Riferimento normativo": p["norma"],
            "N. alunni con organismo": n_ass,
            "Spostamenti in questa fase": n_spost,
        })
    df_passi = pd.DataFrame(righe_passi, columns=[
        "N.", "Fase", "Descrizione", "Riferimento normativo",
        "N. alunni con organismo", "Spostamenti in questa fase",
    ])

    # --- Diario per alunno ---
    intest_passi = [f"{p['indice']}. {p['titolo']}" for p in passi]
    righe_diario = []
    for idx in ordine:
        a = anagrafica[idx]
        riga = {
            "Codice Iscrizione": a.get("codice", ""),
            "Cognome": a.get("cognome", ""),
            "Nome": a.get("nome", ""),
            "Plesso": a.get("plesso", ""),
            "Gruppo 45h": a.get("gruppo_label", ""),
            "Categoria": a.get("categoria", ""),
        }
        n_var = 0
        prev_val = None
        for p, header in zip(passi, intest_passi):
            val = p["stato"].get(idx, "")
            riga[header] = val if val else PLACEHOLDER_NON_ASSEGNATO
            # conta gli spostamenti veri e propri (riassegnazioni o passaggio
            # a "da assegnare"), non la prima assegnazione da vuoto
            if prev_val not in (None, "") and val != prev_val:
                n_var += 1
            prev_val = val
        riga["N. spostamenti"] = n_var
        righe_diario.append(riga)
    df_diario = pd.DataFrame(righe_diario, columns=(
        ["Codice Iscrizione", "Cognome", "Nome", "Plesso", "Gruppo 45h", "Categoria"]
        + intest_passi + ["N. spostamenti"]
    ))

    # --- Movimenti ---
    def _chiave_mov(m):
        a = anagrafica.get(m["idx"], {})
        return (m["passo_indice"], str(a.get("cognome", "")).upper(),
                str(a.get("nome", "")).upper())

    righe_mov = []
    for m in sorted(movimenti, key=_chiave_mov):
        a = anagrafica.get(m["idx"], {})
        righe_mov.append({
            "Fase": f"{m['passo_indice']}. {m['passo_titolo']}",
            "Codice Iscrizione": a.get("codice", ""),
            "Cognome": a.get("cognome", ""),
            "Nome": a.get("nome", ""),
            "Plesso": a.get("plesso", ""),
            "Gruppo 45h": a.get("gruppo_label", ""),
            "Da (organismo)": m["da"] or PLACEHOLDER_NON_ASSEGNATO,
            "A (organismo)": m["a"] or PLACEHOLDER_MANUALE,
            "Motivo": m["motivo"],
        })
    df_movimenti = pd.DataFrame(righe_mov, columns=[
        "Fase", "Codice Iscrizione", "Cognome", "Nome", "Plesso",
        "Gruppo 45h", "Da (organismo)", "A (organismo)", "Motivo",
    ])

    return df_passi, df_diario, df_movimenti, intest_passi


def _riga_intestazione_meta(meta: dict, parametri: dict) -> list[str]:
    """Righe descrittive comuni a Excel e PDF della cronologia."""
    righe = []
    if meta.get("municipio"):
        righe.append(f"Municipio: {meta['municipio']}")
    if meta.get("anno"):
        righe.append(f"Anno scolastico: {meta['anno']}")
    if meta.get("data_elaborazione"):
        righe.append(f"Estrazione MESIS del: {meta['data_elaborazione']}")
    righe.append(f"Documento generato il: {datetime.date.today().strftime('%d/%m/%Y')}")
    modalita = ("in corso d'anno (finestra di attivazione)"
                if parametri.get("corso_anno") else "inizio anno scolastico")
    righe.append(f"Modalità di assegnazione: {modalita}")
    righe.append(f"Soglia minima ore settimanali per organismo: {parametri.get('soglia', 45)}")
    righe.append(
        f"Iterazioni eseguite: {parametri.get('iterazioni', 0)} — "
        f"spostamenti dell'algoritmo: {parametri.get('spostamenti', 0)}"
    )
    if parametri.get("rettifiche_manuali"):
        righe.append(
            f"Rettifiche manuali dell'ufficio: {parametri['rettifiche_manuali']}"
        )
    return righe


def _conteggi_movimenti(cronologia: dict, parametri: dict) -> dict:
    """Restituisce una copia di `parametri` con i conteggi degli spostamenti
    derivati dai movimenti effettivamente presenti nella cronologia, così i
    documenti restano coerenti anche dopo eventuali rettifiche manuali."""
    movs = cronologia.get("movimenti", [])
    n_manual = sum(
        1 for m in movs if m.get("passo_titolo") == "Rettifica manuale dell'ufficio"
    )
    par = dict(parametri)
    par["spostamenti"] = len(movs) - n_manual
    par["rettifiche_manuali"] = n_manual
    return par


def genera_excel_cronologia(
    cronologia: dict, log_lines: list, meta: dict, parametri: dict,
) -> bytes:
    """Foglio di calcolo con l'intera cronologia del procedimento:
    frontespizio, sintesi delle fasi, diario per alunno (variazioni
    evidenziate) e registro dei movimenti."""
    from openpyxl import load_workbook
    from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    parametri = _conteggi_movimenti(cronologia, parametri)

    def _pulisci(v):
        # rimuove i caratteri di controllo che openpyxl rifiuta
        if isinstance(v, str):
            return ILLEGAL_CHARACTERS_RE.sub("", v)
        return v

    def _pulisci_df(df):
        return df.apply(lambda col: col.map(_pulisci)) if not df.empty else df

    BRAND = "6E1626"
    BRAND_LIGHT = "F1DDE1"
    ZEBRA = "F7F3F1"
    CHANGE = "FCEBD2"  # evidenzia le variazioni nel diario
    WHITE = "FFFFFF"

    HDR_FONT = Font(bold=True, color=WHITE, size=10)
    HDR_FILL = PatternFill(start_color=BRAND, end_color=BRAND, fill_type="solid")
    ZEBRA_FILL = PatternFill(start_color=ZEBRA, end_color=ZEBRA, fill_type="solid")
    CHANGE_FILL = PatternFill(start_color=CHANGE, end_color=CHANGE, fill_type="solid")
    TITLE_FONT = Font(bold=True, size=15, color=BRAND)
    SUB_FONT = Font(bold=True, size=11, color=BRAND)
    INFO_FONT = Font(size=10, color="444444")
    THIN = Border(*(4 * (Side(style="thin", color="D8C3C8"),)))

    df_passi, df_diario, df_movimenti, intest_passi = costruisci_cronologia_dfs(cronologia)
    df_passi = _pulisci_df(df_passi)
    df_diario = _pulisci_df(df_diario)
    df_movimenti = _pulisci_df(df_movimenti)
    log_lines = [_pulisci(str(x)) for x in (log_lines or [])]

    def _fmt(ws, n_cols, start_row=1, wrap_last=False):
        for ci in range(1, n_cols + 1):
            c = ws.cell(row=start_row, column=ci)
            c.font = HDR_FONT
            c.fill = HDR_FILL
            c.border = THIN
            c.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
        ws.row_dimensions[start_row].height = 30
        for ri in range(start_row + 1, ws.max_row + 1):
            zebra = (ri - start_row) % 2 == 0
            for ci in range(1, n_cols + 1):
                cell = ws.cell(row=ri, column=ci)
                cell.border = THIN
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if zebra:
                    cell.fill = ZEBRA_FILL
        for ci in range(1, n_cols + 1):
            width = 12
            for ri in range(start_row, min(ws.max_row + 1, 400)):
                v = ws.cell(row=ri, column=ci).value
                if v is not None:
                    width = max(width, min(len(str(v)) + 3, 48))
            ws.column_dimensions[get_column_letter(ci)].width = width
        ws.freeze_panes = ws.cell(row=start_row + 1, column=1).coordinate

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # frontespizio segnaposto (riempito dopo)
        pd.DataFrame({"": [""]}).to_excel(
            writer, sheet_name="Frontespizio", index=False, header=False)
        df_passi.to_excel(writer, sheet_name="Sintesi passaggi", index=False)
        df_diario.to_excel(writer, sheet_name="Diario per alunno", index=False)
        if not df_movimenti.empty:
            df_movimenti.to_excel(writer, sheet_name="Movimenti", index=False)
        else:
            pd.DataFrame({"Esito": [
                "Non si è reso necessario alcuno spostamento: tutte le nuove "
                "iscrizioni sono state assegnate alla 1ª preferenza nel "
                "rispetto del vincolo delle 45 ore settimanali."
            ]}).to_excel(writer, sheet_name="Movimenti", index=False)
        pd.DataFrame({"Registro elaborazione": log_lines or ["(nessuna voce)"]}).to_excel(
            writer, sheet_name="Registro elaborazione", index=False)

    output.seek(0)
    wb = load_workbook(output)

    # --- Frontespizio ---
    ws = wb["Frontespizio"]
    ws["A1"] = "CRONOLOGIA DEL PROCEDIMENTO DI ASSEGNAZIONE OEPAC"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "Servizio OEPAC — Roma Capitale"
    ws["A2"].font = SUB_FONT
    r = 4
    for riga in _riga_intestazione_meta(meta, parametri):
        ws.cell(row=r, column=1, value=_pulisci(riga)).font = INFO_FONT
        r += 1
    r += 1
    ws.cell(row=r, column=1, value="Riferimenti normativi").font = SUB_FONT
    r += 1
    for nota in [
        "Linee Guida approvate con DGC Roma Capitale n. 260/2024 (Art. 5, "
        "commi 5 e 6): vincolo di 45 ore settimanali per organismo e "
        "continuità educativa.",
        "L. 241/1990 (trasparenza del procedimento amministrativo) e "
        "Reg. UE 2016/679, art. 22 (decisione automatizzata): il presente "
        "documento ricostruisce i passaggi del procedimento a fini di "
        "accesso agli atti.",
        "I dati riguardano alunni con disabilità (art. 9 Reg. UE 2016/679) "
        "e sono soggetti a riservatezza.",
    ]:
        c = ws.cell(row=r, column=1, value="• " + nota)
        c.font = INFO_FONT
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 30
        r += 1
    r += 1
    ws.cell(row=r, column=1, value="Contenuto del documento").font = SUB_FONT
    r += 1
    for nota in [
        "Foglio «Sintesi passaggi»: le fasi del procedimento in ordine, con "
        "il numero di variazioni introdotte da ciascuna.",
        "Foglio «Diario per alunno»: l'organismo assegnato ad ogni alunno in "
        "ciascuna fase (il «prima e dopo» progressivo); le celle evidenziate "
        "segnalano un cambiamento rispetto alla fase precedente.",
        "Foglio «Movimenti»: l'elenco puntuale di ogni spostamento, con la "
        "relativa motivazione.",
        "Foglio «Registro elaborazione»: il registro tecnico dell'algoritmo.",
    ]:
        c = ws.cell(row=r, column=1, value="• " + nota)
        c.font = INFO_FONT
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 28
        r += 1
    ws.column_dimensions["A"].width = 118

    # --- Sintesi passaggi ---
    _fmt(wb["Sintesi passaggi"], df_passi.shape[1])

    # --- Diario per alunno: evidenzia solo gli spostamenti reali ---
    # (riassegnazione da un organismo a un altro, o ritorno a non assegnato),
    # non la prima assegnazione da vuoto: coerente con "N. spostamenti".
    ws_d = wb["Diario per alunno"]
    _fmt(ws_d, df_diario.shape[1])
    col_passi_idx = [
        list(df_diario.columns).index(h) + 1 for h in intest_passi
    ]
    for ri in range(2, ws_d.max_row + 1):
        for pos, ci in enumerate(col_passi_idx):
            if pos == 0:
                continue
            val = ws_d.cell(row=ri, column=ci).value
            prev = ws_d.cell(row=ri, column=col_passi_idx[pos - 1]).value
            if prev and prev != PLACEHOLDER_NON_ASSEGNATO and val != prev:
                ws_d.cell(row=ri, column=ci).fill = CHANGE_FILL

    # --- Movimenti ---
    ws_m = wb["Movimenti"]
    _fmt(ws_m, ws_m.max_column)
    if not df_movimenti.empty:
        motivo_idx = list(df_movimenti.columns).index("Motivo") + 1
        ws_m.column_dimensions[get_column_letter(motivo_idx)].width = 70

    # --- Registro ---
    ws_r = wb["Registro elaborazione"]
    _fmt(ws_r, 1)
    ws_r.column_dimensions["A"].width = 100

    final = io.BytesIO()
    wb.save(final)
    final.seek(0)
    return final.getvalue()


def genera_pdf_verbale(
    cronologia: dict, stats: dict, log_lines: list, meta: dict, parametri: dict,
) -> bytes:
    """Verbale formale del procedimento automatizzato, adatto ad essere
    allegato a una risposta di accesso agli atti."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    )

    BRAND = colors.HexColor("#6E1626")
    BRAND_LIGHT = colors.HexColor("#F1DDE1")
    ZEBRA = colors.HexColor("#F7F3F1")

    parametri = _conteggi_movimenti(cronologia, parametri)

    def _xml(s, maxlen=800):
        # escappa &, <, > per l'interprete di markup di reportlab e limita
        # la lunghezza per evitare celle più alte di una pagina
        s = str(s)
        if len(s) > maxlen:
            s = s[:maxlen].rstrip() + "…"
        return html.escape(s, quote=False)

    df_passi, _, df_movimenti, _ = costruisci_cronologia_dfs(cronologia)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=16 * mm, rightMargin=16 * mm,
        topMargin=16 * mm, bottomMargin=14 * mm,
    )
    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("h1", parent=styles["Title"], fontSize=15,
                        textColor=BRAND, spaceAfter=4, leading=19)
    h2 = ParagraphStyle("h2", parent=styles["Heading2"], fontSize=11.5,
                        textColor=BRAND, spaceBefore=10, spaceAfter=5)
    normal = ParagraphStyle("n", parent=styles["Normal"], fontSize=9.2, leading=13)
    small = ParagraphStyle("s", parent=styles["Normal"], fontSize=8,
                           textColor=colors.HexColor("#666666"), leading=11)
    cell = ParagraphStyle("c", parent=styles["Normal"], fontSize=7.6, leading=9.5)
    cellb = ParagraphStyle("cb", parent=cell, fontName="Helvetica-Bold")

    elems = []
    elems.append(Paragraph(
        "VERBALE DEL PROCEDIMENTO DI ASSEGNAZIONE AUTOMATICA<br/>"
        "DEL SERVIZIO OEPAC", h1))
    elems.append(Paragraph("Roma Capitale — Servizio OEPAC", normal))
    elems.append(Spacer(1, 6))

    for riga in _riga_intestazione_meta(meta, parametri):
        elems.append(Paragraph(_xml(riga), normal))
    elems.append(Spacer(1, 8))

    elems.append(Paragraph("Premessa e riferimenti normativi", h2))
    elems.append(Paragraph(
        "Il presente verbale ricostruisce, ai fini della trasparenza "
        "amministrativa (L. 241/1990) e del diritto di accesso agli atti, i "
        "passaggi del procedimento di assegnazione degli alunni con "
        "disabilità agli Organismi accreditati per il servizio OEPAC. "
        "L'assegnazione è effettuata in modo automatizzato secondo le Linee "
        "Guida approvate con DGC Roma Capitale n. 260/2024 (Art. 5, commi 5 e "
        "6): ogni Organismo deve raggiungere di norma almeno 45 ore "
        "settimanali per gruppo (Istituto Comprensivo o ambito territoriale), "
        "mentre le riconferme e gli alunni già attivati sono mantenuti sul "
        "medesimo Organismo per continuità educativa. Ai sensi dell'art. 22 "
        "del Reg. UE 2016/679, si dà evidenza della logica del trattamento "
        "automatizzato.", normal))

    # --- Fasi del procedimento ---
    elems.append(Paragraph("Fasi del procedimento", h2))
    dati_fasi = [[
        Paragraph("<b>N.</b>", cell), Paragraph("<b>Fase</b>", cell),
        Paragraph("<b>Riferimento</b>", cell),
        Paragraph("<b>Alunni con<br/>organismo</b>", cell),
        Paragraph("<b>Spostamenti<br/>nella fase</b>", cell),
    ]]
    for _, rp in df_passi.iterrows():
        dati_fasi.append([
            Paragraph(str(rp["N."]), cell),
            Paragraph(f"<b>{_xml(rp['Fase'])}</b><br/>{_xml(rp['Descrizione'])}", cell),
            Paragraph(_xml(rp["Riferimento normativo"]), cell),
            Paragraph(str(rp["N. alunni con organismo"]), cell),
            Paragraph(str(rp["Spostamenti in questa fase"]), cell),
        ])
    t_fasi = Table(dati_fasi, colWidths=[9 * mm, 81 * mm, 43 * mm, 20 * mm, 25 * mm],
                   repeatRows=1)
    t_fasi.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), BRAND),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D8C3C8")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ALIGN", (3, 0), (-1, -1), "CENTER"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, ZEBRA]),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    elems.append(t_fasi)

    # --- Esiti complessivi ---
    elems.append(Paragraph("Esiti complessivi", h2))
    n_tot = stats.get("totale_alunni", 0)
    n_crit = stats.get("criticita", 0)
    esiti = (
        f"Alunni elaborati: <b>{n_tot}</b> &nbsp;|&nbsp; "
        f"riconferme: <b>{stats.get('riconferme', 0)}</b> &nbsp;|&nbsp; "
        f"nuove iscrizioni: <b>{stats.get('nuove_iscrizioni', 0)}</b>"
    )
    if parametri.get("corso_anno"):
        esiti += f" &nbsp;|&nbsp; già attivati mantenuti: <b>{stats.get('gia_attivati', 0)}</b>"
    esiti += (
        f"<br/>Spostamenti operati dall'algoritmo: "
        f"<b>{parametri.get('spostamenti', 0)}</b> in "
        f"<b>{stats.get('iterazioni', 0)}</b> iterazioni."
    )
    if parametri.get("rettifiche_manuali"):
        esiti += (
            f" &nbsp;|&nbsp; rettifiche manuali dell'ufficio: "
            f"<b>{parametri['rettifiche_manuali']}</b>."
        )
    esiti += f" &nbsp;|&nbsp; Casi da verificare: <b>{n_crit}</b>."
    elems.append(Paragraph(esiti, normal))

    # --- Elenco movimenti ---
    elems.append(Paragraph("Registro degli spostamenti", h2))
    if df_movimenti.empty:
        elems.append(Paragraph(
            "Non si è reso necessario alcuno spostamento: tutte le nuove "
            "iscrizioni sono state assegnate alla 1ª preferenza espressa "
            "nel rispetto del vincolo delle 45 ore settimanali.", normal))
    else:
        dett = f"{parametri.get('spostamenti', 0)} dell'algoritmo"
        if parametri.get("rettifiche_manuali"):
            dett += f", {parametri['rettifiche_manuali']} rettifiche manuali dell'ufficio"
        elems.append(Paragraph(
            f"Sono registrati {len(df_movimenti)} spostamenti ({dett}). Per "
            "ciascuno si riporta l'alunno, il gruppo, l'organismo di partenza "
            "e di destinazione e la motivazione.", normal))
        elems.append(Spacer(1, 4))
        dati_mov = [[
            Paragraph("<b>Fase</b>", cell), Paragraph("<b>Alunno</b>", cell),
            Paragraph("<b>Gruppo 45h</b>", cell),
            Paragraph("<b>Da → A</b>", cell), Paragraph("<b>Motivo</b>", cell),
        ]]
        for _, rm in df_movimenti.iterrows():
            passo_n = str(rm["Fase"]).split(".", 1)[0]
            dati_mov.append([
                Paragraph(passo_n, cell),
                Paragraph(f"{_xml(rm['Cognome'])} {_xml(rm['Nome'])}<br/>"
                          f"<font size=6>{_xml(rm['Plesso'])}</font>", cell),
                Paragraph(_xml(rm["Gruppo 45h"]), cell),
                Paragraph(f"{_xml(rm['Da (organismo)'])} <b>→</b> "
                          f"{_xml(rm['A (organismo)'])}", cell),
                Paragraph(_xml(rm["Motivo"]), cell),
            ])
        t_mov = Table(
            dati_mov,
            colWidths=[12 * mm, 33 * mm, 29 * mm, 46 * mm, 58 * mm],
            repeatRows=1,
        )
        t_mov.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), BRAND),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D8C3C8")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, ZEBRA]),
            ("TOPPADDING", (0, 0), (-1, -1), 2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ]))
        elems.append(t_mov)

    elems.append(Spacer(1, 14))
    elems.append(Paragraph(
        "Il presente documento è generato automaticamente dal sistema di "
        "assegnazione OEPAC e riproduce fedelmente i passaggi del "
        "procedimento. Data ________________________ &nbsp;&nbsp;&nbsp; "
        "Il Responsabile del procedimento ________________________", small))

    doc.build(elems)
    buf.seek(0)
    return buf.getvalue()


def registra_rettifica_manuale(cronologia: dict, modifiche: list[dict]):
    """Aggiunge alla cronologia una fase che documenta le rettifiche manuali
    operate dall'ufficio dopo l'assegnazione automatica (trasparenza).

    `modifiche`: lista di dict con chiavi codice, cognome, nome, plesso,
    gruppo_label, da, a.
    """
    if not cronologia or not modifiche:
        return
    passi = cronologia.setdefault("passi", [])
    movimenti = cronologia.setdefault("movimenti", [])
    anagrafica = cronologia.setdefault("anagrafica", {})

    # stato di partenza = ultimo passo registrato
    stato_prec = dict(passi[-1]["stato"]) if passi else {}
    # indicizza l'anagrafica per codice per ritrovare l'idx interno
    idx_per_codice = {
        str(a.get("codice", "")): i for i, a in anagrafica.items()
    }
    nuovo_stato = dict(stato_prec)
    passo_idx = len(passi)
    titolo = "Rettifica manuale dell'ufficio"
    for m in modifiche:
        cod = str(m.get("codice", ""))
        idx = idx_per_codice.get(cod)
        a = m.get("a", "") or ""
        if idx is not None:
            nuovo_stato[idx] = a.strip() if isinstance(a, str) else ""
        else:
            # alunno non presente nell'anagrafica interna: lo aggiunge
            idx = f"manuale-{cod}"
            anagrafica[idx] = {
                "codice": cod, "cognome": m.get("cognome", ""),
                "nome": m.get("nome", ""), "plesso": m.get("plesso", ""),
                "gruppo": "", "gruppo_label": m.get("gruppo_label", ""),
                "ore": m.get("ore", 0), "categoria": "Rettifica manuale",
            }
            for p in passi:
                p["stato"].setdefault(idx, "")
            stato_prec.setdefault(idx, "")
            nuovo_stato[idx] = a.strip() if isinstance(a, str) else ""
        movimenti.append({
            "idx": idx, "passo_indice": passo_idx, "passo_titolo": titolo,
            "da": m.get("da", "") or "", "a": m.get("a", "") or "",
            "motivo": "Rettifica manuale disposta dall'ufficio dopo "
                      "l'assegnazione automatica.",
        })
    passi.append({
        "indice": passo_idx,
        "titolo": titolo,
        "descrizione": f"{len(modifiche)} assegnazioni modificate manualmente "
                       "dall'ufficio rispetto all'esito automatico.",
        "norma": "Intervento manuale dell'ufficio",
        "stato": nuovo_stato,
    })


# ---------------------------------------------------------------------------
# Consistency checks
# ---------------------------------------------------------------------------

def verifiche_consistenza(
    df_result, df_input_work, col_map, stats, auto_ambito=False,
) -> list[tuple[str, bool, str]]:
    checks = []

    checks.append((
        "Alunni processati",
        True,
        f"{stats['totale_alunni']} alunni con Stato=ATTIVA e senza rinuncia",
    ))

    all_have_status = df_result["Status"].notna().all() and (df_result["Status"] != "").all()
    checks.append((
        "Tutti gli alunni hanno status valorizzato",
        all_have_status,
        "OK" if all_have_status else "ERRORE: alcuni alunni senza status",
    ))

    riconferme_spostate = df_result.loc[
        (df_result["Tipo"].str.strip().str.upper() == "RICONFERMA")
        & (df_result["Organismo Assegnato dall'Algoritmo"] != "")
        & (df_result["Organismo Pre-esistente"] != "")
        & (df_result["Organismo Assegnato dall'Algoritmo"] != df_result["Organismo Pre-esistente"])
    ]
    ok_ric = len(riconferme_spostate) == 0
    checks.append((
        "Riconferme mantenute (nessuna spostata d'ufficio)",
        ok_ric,
        "OK" if ok_ric else f"ATTENZIONE: {len(riconferme_spostate)} riconferme con organismo diverso",
    ))

    if "Preferenza Soddisfatta" in df_result.columns:
        attivati_spostati = df_result.loc[
            (df_result["Preferenza Soddisfatta"] == "Già attivato")
            & (df_result["Organismo Assegnato dall'Algoritmo"] != df_result["Organismo Pre-esistente"])
        ]
        ok_att = len(attivati_spostati) == 0
        if (df_result["Preferenza Soddisfatta"] == "Già attivato").any():
            checks.append((
                "Alunni già attivati mantenuti sul loro organismo",
                ok_att,
                "OK" if ok_att else f"ERRORE: {len(attivati_spostati)} alunni attivati spostati",
            ))

    if auto_ambito and "Tipo Gestione" in df_result.columns:
        plessi_nd = df_result.loc[
            df_result["Tipo Gestione"] == "N/D", "Codice Meccanografico Plesso"
        ].nunique()
        checks.append((
            "Classificazione automatica completa",
            plessi_nd == 0,
            "Tutti i plessi classificati (Statale / Infanzia comunale / Paritaria)"
            if plessi_nd == 0
            else f"ATTENZIONE: {plessi_nd} plessi non classificati (N/D) — gruppo derivato dal solo plesso",
        ))

        presunti = df_result.loc[
            df_result["Fonte Classificazione"] == "presunta",
            "Codice Meccanografico Plesso",
        ].nunique()
        checks.append((
            "Classificazioni da verificare",
            presunti == 0,
            "Nessuna classificazione presunta"
            if presunti == 0
            else f"ATTENZIONE: {presunti} plessi con classificazione presunta — "
                 "verificare nella tabella 'Classificazione scuole'",
        ))

        mask_mun = df_result["Tipo Gestione"].isin([TIPO_COMUNALE, TIPO_PARITARIA])
        senza_mun = df_result.loc[
            mask_mun
            & (df_result["Municipio"] == "")
            & (~df_result["Gruppo 45h"].str.contains("Ambito", na=False)),
            "Codice Meccanografico Plesso",
        ].nunique()
        checks.append((
            "Ambito/municipio riconosciuto per comunali e paritarie",
            senza_mun == 0,
            "OK" if senza_mun == 0
            else f"ATTENZIONE: {senza_mun} plessi senza ambito né municipio riconosciuto",
        ))

        amb_up = df_result["Ambito"].fillna("").str.upper()
        conflitti = df_result.loc[
            (amb_up.str.contains("PARIT") & (df_result["Tipo Gestione"] == TIPO_COMUNALE))
            | (amb_up.str.contains("COMUNAL") & (df_result["Tipo Gestione"] == TIPO_PARITARIA)),
            "Codice Meccanografico Plesso",
        ].nunique()
        checks.append((
            "Coerenza con la colonna Ambito del MESIS",
            conflitti == 0,
            "Nessun conflitto" if conflitti == 0
            else f"ATTENZIONE: {conflitti} plessi con classificazione in conflitto "
                 "con la colonna Ambito — verificare",
        ))

        # la numerazione degli ambiti territoriali è cittadina (1..37):
        # lo stesso ambito in più municipi indica dati da verificare
        df_amb = df_result.loc[
            df_result["Gruppo 45h"].str.contains("Ambito", na=False)
        ]
        multi_mun = []
        for grp, df_g in df_amb.groupby("Gruppo 45h"):
            municipi = {m for m in df_g["Municipio"] if m}
            if len(municipi) > 1:
                multi_mun.append(f"{grp} ({', '.join(sorted(municipi))})")
        checks.append((
            "Ambiti territoriali univoci tra municipi",
            not multi_mun,
            "OK" if not multi_mun
            else "ATTENZIONE: gruppi con scuole di municipi diversi: "
                 + "; ".join(multi_mun) + " — verificare la colonna Ambito",
        ))

    ore_output = df_result["Ore Assegnate"].sum() if len(df_result) > 0 else 0
    ore_col = col_map.get("ore_assegnate")
    stato_col = col_map.get("stato")
    rinuncia_col = col_map.get("data_rinuncia")
    if ore_col and stato_col:
        mask_in = df_input_work[stato_col].str.strip().str.upper() == "ATTIVA"
        if rinuncia_col:
            mask_in &= df_input_work[rinuncia_col].isna() | (df_input_work[rinuncia_col].str.strip() == "")
        ore_input = pd.to_numeric(
            df_input_work.loc[mask_in, ore_col], errors="coerce"
        ).fillna(0).sum()
        ore_ok = abs(ore_output - ore_input) < 0.01
        checks.append((
            "Somma ore coerente",
            ore_ok,
            f"Input: {ore_input} — Output: {ore_output}"
            + ("" if ore_ok else " — DISALLINEAMENTO"),
        ))
    else:
        checks.append((
            "Somma ore coerente",
            True,
            f"Ore totali output: {ore_output}",
        ))

    return checks


# ---------------------------------------------------------------------------
# Streamlit UI
# ---------------------------------------------------------------------------

STILE_APP = """
<style>
#MainMenu, footer {visibility: hidden;}
div[data-testid="stDecoration"] {display: none;}

.oepac-header {
    background: linear-gradient(135deg, #6E1626 0%, #9E2B3D 100%);
    border-radius: 12px;
    padding: 20px 26px;
    margin-bottom: 4px;
    color: #FFFFFF;
}
.oepac-header .titolo {
    font-size: 1.55rem; font-weight: 700; letter-spacing: .2px;
    line-height: 1.25;
}
.oepac-header .sottotitolo {
    font-size: .9rem; opacity: .92; margin-top: 3px;
}

.oepac-steps { display: flex; gap: 8px; margin: 14px 0 6px 0; flex-wrap: wrap; }
.oepac-step {
    flex: 1 1 170px; display: flex; align-items: center; gap: 9px;
    padding: 9px 12px; border-radius: 9px; font-size: .85rem;
    background: #F4F1EE; color: #6B7280; border: 1px solid #E7E2DC;
}
.oepac-step .n {
    display: inline-flex; align-items: center; justify-content: center;
    width: 22px; height: 22px; border-radius: 50%;
    background: #C9C2BA; color: #FFF; font-weight: 700; font-size: .78rem;
    flex: 0 0 22px;
}
.oepac-step.attivo {
    background: #FBF3F0; color: #6E1626; border-color: #E2C3BC; font-weight: 600;
}
.oepac-step.attivo .n { background: #6E1626; }
.oepac-step.fatto { background: #F1F7F2; color: #2F6846; border-color: #CDE3D4; }
.oepac-step.fatto .n { background: #3D8B5F; }

div[data-testid="stMetric"] {
    background: #FFFFFF; border: 1px solid #E7E2DC; border-radius: 10px;
    padding: 12px 16px; box-shadow: 0 1px 3px rgba(60, 40, 30, .05);
}
div[data-testid="stMetric"] label { color: #6B7280; }
</style>
"""

PASSI_APP = [
    "Carica il file MESIS",
    "Controlla dati e scuole",
    "Esegui l'assegnazione",
    "Scarica i risultati",
]


def _render_intestazione():
    st.markdown(STILE_APP, unsafe_allow_html=True)
    st.markdown(
        """
        <div class="oepac-header">
          <div class="titolo">Assegnazione automatica OEPAC</div>
          <div class="sottotitolo">Roma Capitale · Linee Guida DGC n. 260/2024
          (Art. 5, commi 5 e 6) · vincolo 45 ore settimanali per gruppo ·
          classificazione automatica IC / infanzie comunali / scuole paritarie</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def _render_passi(corrente: int):
    """Indicatore del percorso guidato (1..4)."""
    blocchi = []
    for i, nome in enumerate(PASSI_APP, 1):
        if i < corrente:
            cls, segno = "fatto", "&#10003;"
        elif i == corrente:
            cls, segno = "attivo", str(i)
        else:
            cls, segno = "", str(i)
        blocchi.append(
            f'<div class="oepac-step {cls}"><span class="n">{segno}</span>{nome}</div>'
        )
    st.markdown(
        f'<div class="oepac-steps">{"".join(blocchi)}</div>',
        unsafe_allow_html=True,
    )


def _render_benvenuto():
    """Pagina iniziale con le istruzioni, mostrata finché non c'è un file."""
    st.markdown("##### Come funziona")
    descrizioni = [
        ("1 · Carica", "Esporta da MESIS l'elenco iscrizioni OEPAC del tuo "
         "municipio in formato Excel (.xlsx) e caricalo qui sopra."),
        ("2 · Controlla", "Verifica i numeri dell'estrazione e la "
         "classificazione automatica delle scuole: Istituti Comprensivi, "
         "infanzie comunali e paritarie, con i rispettivi gruppi 45h."),
        ("3 · Esegui", "L'algoritmo assegna le nuove domande rispettando le "
         "preferenze delle famiglie, la continuità educativa e il vincolo "
         "delle 45 ore settimanali per gruppo."),
        ("4 · Scarica", "Report Excel completo, un file riservato per ogni "
         "organismo e le lettere di assegnazione in PDF, pronti per l'invio."),
    ]
    cols = st.columns(4)
    for col, (titolo, testo) in zip(cols, descrizioni):
        with col, st.container(border=True):
            st.markdown(f"**{titolo}**")
            st.caption(testo)

    c_sx, c_dx = st.columns(2)
    with c_sx, st.expander("Le quattro finestre di attivazione"):
        st.markdown(
            "Indicazioni operative del Dipartimento Scuola "
            "(nota QM/102670/2025):\n\n"
            "| Finestra | Domande presentate | Attivazione servizio |\n"
            "|---|---|---|\n"
            "| 1ª | entro il 15 luglio | da inizio anno scolastico |\n"
            "| 2ª | 16 luglio – 15 ottobre | da novembre |\n"
            "| 3ª | 16 ottobre – 15 gennaio | da febbraio |\n"
            "| 4ª | 16 gennaio – 15 marzo | da aprile |\n\n"
            "Per le finestre in corso d'anno seleziona la modalità "
            "**In corso d'anno** nelle impostazioni a sinistra: gli alunni "
            "già attivati restano sul loro organismo e vengono assegnate "
            "solo le nuove domande."
        )
    with c_dx, st.expander("Requisiti del file MESIS"):
        st.markdown(
            "- Formato **.xlsx** esportato da MESIS (elenco iscrizioni OEPAC)\n"
            "- Righe 1-7: metadati dell'estrazione (vengono ignorati)\n"
            "- Riga 8: intestazioni delle colonne\n"
            "- Dalla riga 9: un alunno per riga\n\n"
            "Colonne necessarie: Codice, Stato, Tipo, Cognome, Nome, "
            "Codice Meccanografico Plesso, Ore Assegnate. Le altre colonne "
            "(Municipio, Ambito, Istituto, preferenze…) migliorano la "
            "classificazione automatica e i report."
        )


def estrai_metadati_mesis(file_bytes: bytes) -> dict[str, str]:
    """Legge le righe di metadati in testa al file MESIS (anno scolastico,
    municipio, data di elaborazione)."""
    info: dict[str, str] = {}
    try:
        df_meta = pd.read_excel(
            io.BytesIO(file_bytes), sheet_name=0, header=None, nrows=7, dtype=str,
        )
        for v in df_meta[0].dropna():
            s = str(v).strip()
            low = s.lower()
            if ":" not in s:
                continue
            valore = s.split(":", 1)[1].strip().strip(",")
            if low.startswith("anno scolastico"):
                info["anno"] = valore
            elif low.startswith("municipio"):
                info["municipio"] = valore
            elif low.startswith("data elaborazione"):
                info["data_elaborazione"] = valore
    except Exception:
        pass
    return info


def main():
    st.set_page_config(
        page_title="Assegnazione OEPAC — Roma Capitale",
        page_icon="🏛️",
        layout="wide",
        initial_sidebar_state="expanded",
    )

    _render_intestazione()

    with st.sidebar:
        st.markdown("### Impostazioni")

        modalita = st.radio(
            "Momento dell'assegnazione",
            ["Inizio anno scolastico", "In corso d'anno (finestra di attivazione)"],
            help=(
                "Inizio anno: le nuove iscrizioni vengono assegnate dalle "
                "preferenze con il vincolo 45h. In corso d'anno: gli alunni "
                "già attivati (Organismo Assegnato e Data Attivazione "
                "presenti nel MESIS) restano sul loro organismo per "
                "continuità e l'algoritmo assegna solo le nuove domande "
                "della finestra di attivazione. Le finestre OEPAC (nota "
                "QM/102670/2025) sono quattro: domande entro il 15/07 -> "
                "attivazione da inizio anno; 16/07-15/10 -> da novembre; "
                "16/10-15/01 -> da febbraio; 16/01-15/03 -> da aprile."
            ),
        )
        corso_anno = modalita.startswith("In corso")
        data_riferimento = None
        if corso_anno:
            usa_data_rif = st.checkbox(
                "Considera attivati solo prima di una data",
                value=False,
                help=(
                    "Se attivo, solo gli alunni con Data Attivazione fino "
                    "alla data indicata sono considerati già attivati; gli "
                    "altri vengono riassegnati dall'algoritmo."
                ),
            )
            if usa_data_rif:
                data_riferimento = st.date_input(
                    "Attivati fino al", value=datetime.date.today(),
                    format="DD/MM/YYYY",
                )

        soglia = st.slider(
            "Soglia minima ore settimanali",
            min_value=0, max_value=100, value=45,
            help=(
                "Monte ore settimanale che ogni Organismo deve raggiungere "
                "nel gruppo (IC o Ambito) per l'affidamento — di norma 45 "
                "(Art. 5, comma 5). Modificare solo su indicazione della "
                "Direzione Socio-Educativa."
            ),
        )

        auto_ambito = st.toggle(
            "Classificazione automatica ambiti (Roma)",
            value=True,
            help=(
                "Deriva automaticamente i gruppi 45h dal codice "
                "meccanografico e dai dati MESIS: un gruppo per ogni "
                "Istituto Comprensivo statale, e gruppi separati per le "
                "infanzie comunali e per le scuole paritarie (per ambito "
                "territoriale, o per municipio se l'ambito manca dal file). "
                "Se disattivo, i gruppi seguono la sola colonna Ambito "
                "del MESIS."
            ),
        )

        with st.expander("Parametri economici"):
            n_settimane = st.number_input(
                "Settimane annuali", min_value=1, max_value=52, value=35,
                help="Durata della convenzione annuale (di norma 35 settimane).",
            )
            perc_decurtazione = st.number_input(
                "Decurtazione %", min_value=0.0, max_value=50.0, value=11.0, step=0.5,
                help="Percentuale di decurtazione applicata alle ore annuali lorde.",
            )
            costo_orario = st.number_input(
                "Costo orario (EUR)", min_value=0.0, value=24.07, step=0.01,
                format="%.2f",
            )
            aliquota_iva = st.number_input(
                "Aliquota IVA %", min_value=0.0, max_value=30.0, value=5.0, step=0.5,
            )

        with st.expander("File opzionali"):
            st.markdown("**Anagrafe plessi**")
            st.caption(
                "Arricchisce l'indirizzo dei plessi nei report. File "
                ".xlsx/.csv con colonne codice meccanografico e indirizzo "
                "(scaricabile da dati.istruzione.it)."
            )
            anagrafe_file = st.file_uploader(
                "File anagrafe",
                type=["xlsx", "csv"],
                key="anagrafe",
                label_visibility="collapsed",
            )
            st.markdown("**Elenco scuole comunali**")
            st.caption(
                "Per una classificazione certa delle infanzie comunali: "
                "file .xlsx/.csv con una colonna di codici meccanografici "
                "delle scuole dell'infanzia comunali di Roma Capitale."
            )
            comunali_file = st.file_uploader(
                "File elenco comunali",
                type=["xlsx", "csv"],
                key="comunali",
                label_visibility="collapsed",
            )

        with st.expander("Impostazioni avanzate"):
            max_iter = st.number_input(
                "Iterazioni massime algoritmo",
                min_value=1, max_value=200, value=50,
                help="Limite di sicurezza per la convergenza dell'algoritmo.",
            )

    passi_slot = st.empty()

    def _mostra_passi(n: int):
        with passi_slot.container():
            _render_passi(n)

    st.markdown("#### 1 · Carica il file MESIS")
    uploaded = st.file_uploader(
        "Trascina qui l'elenco iscrizioni OEPAC esportato da MESIS (.xlsx) "
        "oppure usa *Browse files*",
        type=["xlsx"],
        key="mesis",
        help=(
            "Esportazione MESIS standard: 7 righe di metadati, intestazioni "
            "alla riga 8, un alunno per riga dalla riga 9."
        ),
    )

    if uploaded is None:
        st.session_state.pop("risultati", None)
        st.session_state.pop("_last_file_id", None)
        _mostra_passi(1)
        _render_benvenuto()
        return

    file_id = f"{uploaded.name}_{uploaded.size}"
    if file_id != st.session_state.get("_last_file_id"):
        st.session_state.pop("risultati", None)
        st.session_state["_last_file_id"] = file_id

    file_bytes = uploaded.getvalue()
    df_raw, col_map, errors = carica_dati(file_bytes)

    if errors:
        st.error(
            "Il file caricato non rispetta il formato MESIS atteso. "
            "Controlla di aver esportato l'elenco iscrizioni OEPAC in "
            "formato .xlsx senza modificarne la struttura."
        )
        for err in errors:
            st.warning(err)
        if not col_map or df_raw.empty:
            _mostra_passi(1)
            return

    stato_col = col_map.get("stato")
    rinuncia_col = col_map.get("data_rinuncia")
    tipo_col = col_map.get("tipo")
    ore_col = col_map.get("ore_assegnate")
    mask_attiva = df_raw[stato_col].str.strip().str.upper() == "ATTIVA"
    if rinuncia_col:
        mask_attiva &= df_raw[rinuncia_col].isna() | (df_raw[rinuncia_col].str.strip() == "")
    n_attive = int(mask_attiva.sum())
    n_ric = 0
    n_nuove = 0
    if tipo_col:
        tipo_vals = df_raw.loc[mask_attiva, tipo_col].str.strip().str.upper()
        n_ric = int((tipo_vals == "RICONFERMA").sum())
        n_nuove = int((tipo_vals == "NUOVA ISCRIZIONE").sum())
    ore_tot = 0
    if ore_col:
        ore_tot = pd.to_numeric(df_raw.loc[mask_attiva, ore_col], errors="coerce").fillna(0).sum()

    org_prev_col = col_map.get("organismo_assegnato")
    att_prev_col = col_map.get("data_attivazione")
    n_attivati = 0
    if corso_anno and tipo_col and org_prev_col and att_prev_col:
        sub = df_raw.loc[mask_attiva]
        mask_att_prev = (
            (sub[tipo_col].str.strip().str.upper() == "NUOVA ISCRIZIONE")
            & sub[org_prev_col].fillna("").str.strip().ne("")
            & sub[att_prev_col].fillna("").str.strip().ne("")
        )
        n_attivati = int(mask_att_prev.sum())

    st.markdown("#### 2 · Controlla i dati")
    meta_mesis = estrai_metadati_mesis(file_bytes)
    info_bits = []
    if meta_mesis.get("municipio"):
        info_bits.append(f"**{meta_mesis['municipio'].title()}**")
    if meta_mesis.get("anno"):
        info_bits.append(f"anno scolastico **{meta_mesis['anno']}**")
    if meta_mesis.get("data_elaborazione"):
        info_bits.append(f"estrazione MESIS del {meta_mesis['data_elaborazione']}")
    if info_bits:
        st.markdown("File riconosciuto: " + " · ".join(info_bits))

    if corso_anno:
        c1, c2, c3, c4, c5 = st.columns(5)
        c3.metric("Nuove iscrizioni", f"{n_nuove}")
        c4.metric("di cui già attivate", f"{n_attivati}")
        c5.metric("Ore totali", f"{ore_tot:,.0f}")
    else:
        c1, c2, c3, c4 = st.columns(4)
        c3.metric("Nuove iscrizioni", f"{n_nuove}")
        c4.metric("Ore totali", f"{ore_tot:,.0f}")
    c1.metric("Iscrizioni attive", f"{n_attive}")
    c2.metric("Riconferme", f"{n_ric}")

    codici_comunali: set[str] = set()
    if comunali_file is not None:
        codici_comunali = carica_codici_comunali(
            comunali_file.getvalue(), comunali_file.name,
        )
        if codici_comunali:
            st.sidebar.success(f"Elenco comunali: {len(codici_comunali)} codici caricati.")
        else:
            st.sidebar.warning(
                "Elenco comunali: nessun codice riconosciuto nel file caricato."
            )

    tipi_override: dict[str, str] = {}
    if auto_ambito:
        df_scuole = costruisci_tabella_scuole(df_raw, col_map, codici_comunali)
        if not df_scuole.empty:
            n_presunte = int((df_scuole["Fonte"] == "presunta").sum())
            label_exp = f"Classificazione scuole ({len(df_scuole)} plessi"
            label_exp += f", {n_presunte} da verificare)" if n_presunte else ")"
            with st.expander(label_exp, expanded=n_presunte > 0):
                st.caption(
                    "Classificazione derivata automaticamente da codice "
                    "meccanografico, colonna Ambito e denominazioni. "
                    "Correggi **Tipo Gestione** dove serve: la correzione "
                    "viene applicata ai gruppi 45h alla prossima esecuzione. "
                    "La colonna *Gruppo 45h (auto)* mostra il gruppo prima "
                    "delle correzioni."
                )
                opzioni_tipo = TIPI_GESTIONE + (
                    ["N/D"] if (df_scuole["Tipo Gestione"] == "N/D").any() else []
                )
                edited_scuole = st.data_editor(
                    df_scuole,
                    use_container_width=True,
                    hide_index=True,
                    key=f"editor_scuole_{file_id}",
                    column_config={
                        "Tipo Gestione": st.column_config.SelectboxColumn(
                            "Tipo Gestione",
                            options=opzioni_tipo,
                            required=True,
                        ),
                    },
                    disabled=[c for c in df_scuole.columns if c != "Tipo Gestione"],
                )
                for i in range(len(df_scuole)):
                    cod = df_scuole.iloc[i]["Codice Meccanografico Plesso"]
                    if edited_scuole.iloc[i]["Tipo Gestione"] != df_scuole.iloc[i]["Tipo Gestione"]:
                        tipi_override[cod] = edited_scuole.iloc[i]["Tipo Gestione"]
                if tipi_override:
                    st.info(f"{len(tipi_override)} classificazioni corrette manualmente.")

    organismi_set = set()
    org_col_name = col_map.get("organismo_assegnato")
    enti_col_name = col_map.get("enti_erogatori_scelti")
    if org_col_name:
        organismi_set.update(df_raw[org_col_name].dropna().str.strip().unique())
    if enti_col_name:
        for val in df_raw[enti_col_name].dropna():
            for pref in parse_preferenze(str(val)):
                organismi_set.add(pref.strip())
    organismi_set.discard("")
    organismi_list = sorted(organismi_set)

    filtro_org = "Tutti"
    if organismi_list:
        with st.sidebar:
            st.divider()
            st.markdown("### Filtri risultati")
            filtro_org = st.selectbox(
                "Mostra solo l'organismo",
                ["Tutti"] + organismi_list,
                help=(
                    "Filtra le tabelle dei risultati sugli alunni di un "
                    "singolo organismo. I file scaricati contengono "
                    "sempre tutti i dati."
                ),
            )

    st.markdown("#### 3 · Esegui l'assegnazione")
    col_btn, col_info = st.columns([1, 3])
    with col_btn:
        run_clicked = st.button(
            "Esegui assegnazione", type="primary", use_container_width=True,
            icon=":material/play_arrow:",
        )
    with col_info:
        if corso_anno:
            st.caption(
                f"Modalità **in corso d'anno**: {n_attivati} alunni già attivati "
                f"restano sul loro organismo per continuità; l'algoritmo "
                f"assegnerà le restanti nuove domande applicando il vincolo "
                f"di {soglia} ore per gruppo. Le riconferme non vengono mai "
                "spostate."
            )
        else:
            st.caption(
                f"Modalità **inizio anno**: l'algoritmo assegnerà le {n_nuove} "
                f"nuove iscrizioni in base alle preferenze delle famiglie, "
                f"applicando il vincolo di {soglia} ore per gruppo. Le "
                "riconferme non vengono mai spostate."
            )

    if run_clicked:
        with st.spinner("Elaborazione in corso..."):
            result = esegui_assegnazione(
                df_raw, col_map, soglia, max_iter,
                auto_ambito=auto_ambito,
                tipi_override=tipi_override,
                codici_comunali=codici_comunali,
                corso_anno=corso_anno,
                data_riferimento=data_riferimento,
            )
            df_assegnazioni, df_riepilogo, df_criticita, log_lines, stats, nome_registro = result
            if anagrafe_file:
                df_assegnazioni = arricchisci_indirizzi(
                    df_assegnazioni, anagrafe_file.getvalue(), filename=anagrafe_file.name,
                )
        st.session_state["risultati"] = {
            "df_assegnazioni": df_assegnazioni,
            "df_riepilogo": df_riepilogo,
            "df_criticita": df_criticita,
            "log": log_lines,
            "stats": stats,
            "df_raw": df_raw,
            "col_map": col_map,
            "auto_ambito": auto_ambito,
            "corso_anno": corso_anno,
            "meta_mesis": estrai_metadati_mesis(file_bytes),
            "parametri": {
                "soglia": soglia,
                "corso_anno": corso_anno,
                "max_iter": max_iter,
                "iterazioni": stats.get("iterazioni", 0),
                "spostamenti": stats.get("spostamenti", 0),
            },
        }

    if "risultati" not in st.session_state:
        _mostra_passi(3)
        return

    res = st.session_state["risultati"]
    df_assegnazioni = res["df_assegnazioni"]
    df_riepilogo = res["df_riepilogo"]
    df_criticita = res["df_criticita"]
    log_lines = res["log"]
    stats = res["stats"]
    _mostra_passi(4)

    st.markdown("---")
    st.markdown("#### 4 · Risultati e download")
    st.caption(
        "Risultati calcolati con: modalità **{}** · classificazione "
        "automatica **{}**. Se cambi le impostazioni nella barra laterale, "
        "riesegui l'assegnazione per aggiornarli.".format(
            "in corso d'anno" if res.get("corso_anno") else "inizio anno",
            "attiva" if res.get("auto_ambito") else "disattivata",
        )
    )

    n_crit_tot = stats["criticita"]
    if n_crit_tot == 0:
        st.success(
            f"Assegnazione completata senza criticità: "
            f"{stats['totale_alunni']} alunni elaborati."
        )
    else:
        st.warning(
            f"Assegnazione completata: {stats['totale_alunni']} alunni "
            f"elaborati, di cui **{n_crit_tot} da verificare** — l'elenco "
            "con le azioni suggerite è nella scheda *Criticità*."
        )

    if res.get("corso_anno") and stats.get("gia_attivati", 0) > 0:
        rc0, rc1, rc2, rc3 = st.columns(4)
        rc0.metric("Già attivati (mantenuti)", stats["gia_attivati"])
    else:
        rc1, rc2, rc3 = st.columns(3)
    with rc1:
        st.metric("Spostamenti effettuati", stats["spostamenti"])
    with rc2:
        st.metric("Iterazioni algoritmo", stats["iterazioni"])
    with rc3:
        n_crit = stats["criticita"]
        st.metric("Casi da verificare", n_crit, delta=None if n_crit == 0 else f"{n_crit} criticita", delta_color="off" if n_crit == 0 else "inverse")

    checks = verifiche_consistenza(
        df_assegnazioni, res["df_raw"], res["col_map"], stats,
        auto_ambito=res.get("auto_ambito", False),
    )
    tutte_ok = all(ok for _, ok, _ in checks)
    label_verifiche = (
        "Verifiche di consistenza — tutte superate" if tutte_ok
        else "Verifiche di consistenza — alcune richiedono attenzione"
    )
    with st.expander(label_verifiche, expanded=not tutte_ok):
        for label, ok, detail in checks:
            st.markdown(
                f":{'green' if ok else 'red'}[{'OK' if ok else 'ATTENZIONE'}] "
                f"**{label}** — {detail}"
            )

    df_display = df_assegnazioni.copy()
    if filtro_org != "Tutti":
        mask = (
            df_display["Organismo Assegnato dall'Algoritmo"].apply(
                lambda x: nomi_uguali(x, filtro_org) if isinstance(x, str) else False
            )
            | df_display["Organismo Pre-esistente"].apply(
                lambda x: nomi_uguali(x, filtro_org) if isinstance(x, str) else False
            )
        )
        df_display = df_display.loc[mask]
        df_riep_display = df_riepilogo.loc[
            df_riepilogo["Organismo"].apply(
                lambda x: nomi_uguali(x, filtro_org) if isinstance(x, str) else False
            )
        ]
    else:
        df_riep_display = df_riepilogo

    tab_ass, tab_riep, tab_crit, tab_graf, tab_log = st.tabs([
        f"📋 Assegnazioni ({len(df_display)})",
        f"📊 Riepilogo gruppi ({len(df_riep_display)})",
        f"⚠️ Criticità ({len(df_criticita)})",
        "📈 Grafici",
        "🧾 Log elaborazione",
    ])

    with tab_ass:
        edit_mode = st.toggle(
            "Modifica manuale assegnazioni",
            help="Permette di cambiare l'Organismo assegnato a un alunno e ricalcolare i riepiloghi.",
        )
        if edit_mode:
            st.caption(
                "Modifica la colonna **Organismo Assegnato dall'Algoritmo**, "
                "poi clicca **Applica modifiche** per ricalcolare riepiloghi e importi. "
                "Le modifiche manuali sono evidenziate con Preferenza = 'Manuale'."
            )
            org_options = sorted([o for o in df_assegnazioni["Organismo Assegnato dall'Algoritmo"].unique() if o])
            edited = st.data_editor(
                df_display,
                use_container_width=True,
                hide_index=True,
                height=500,
                key="editor_ass",
                column_config={
                    "Organismo Assegnato dall'Algoritmo": st.column_config.SelectboxColumn(
                        "Organismo Assegnato dall'Algoritmo",
                        options=org_options,
                        required=False,
                    ),
                },
                disabled=[c for c in df_display.columns if c != "Organismo Assegnato dall'Algoritmo"],
            )
            if st.button("Applica modifiche e ricalcola", type="primary"):
                df_full = df_assegnazioni.set_index("Codice Iscrizione")
                ed = edited.set_index("Codice Iscrizione")
                n_mod = 0
                modifiche_log = []
                for cod, row in ed.iterrows():
                    nuovo = row["Organismo Assegnato dall'Algoritmo"]
                    if cod in df_full.index:
                        vecchio = df_full.at[cod, "Organismo Assegnato dall'Algoritmo"]
                        if str(nuovo) != str(vecchio):
                            modifiche_log.append({
                                "codice": cod,
                                "cognome": df_full.at[cod, "Cognome"] if "Cognome" in df_full.columns else "",
                                "nome": df_full.at[cod, "Nome"] if "Nome" in df_full.columns else "",
                                "plesso": df_full.at[cod, "Plesso"] if "Plesso" in df_full.columns else "",
                                "gruppo_label": etichetta_gruppo(
                                    df_full.at[cod, "Gruppo 45h"] if "Gruppo 45h" in df_full.columns else "",
                                    df_full.at[cod, "Istituto"] if "Istituto" in df_full.columns else "",
                                ),
                                "da": vecchio,
                                "a": nuovo,
                            })
                            df_full.at[cod, "Organismo Assegnato dall'Algoritmo"] = nuovo
                            df_full.at[cod, "Preferenza Soddisfatta"] = "Manuale"
                            df_full.at[cod, "Status"] = "OK" if nuovo else "Da assegnare manualmente"
                            df_full.at[cod, "Note"] = "Assegnazione modificata manualmente"
                            n_mod += 1
                df_new = df_full.reset_index()
                # usa la soglia con cui è stata eseguita l'assegnazione, non il
                # valore live dello slider (che potrebbe essere stato spostato)
                soglia_run = res.get("parametri", {}).get("soglia", soglia)
                new_riep = costruisci_riepilogo_gruppo(df_new, soglia_run)
                new_crit = costruisci_criticita(df_new)
                st.session_state["risultati"]["df_assegnazioni"] = df_new
                st.session_state["risultati"]["df_riepilogo"] = new_riep
                st.session_state["risultati"]["df_criticita"] = new_crit
                st.session_state["risultati"]["stats"]["criticita"] = int((df_new["Status"] != "OK").sum())
                # documenta le rettifiche manuali nella cronologia del procedimento
                cronologia = st.session_state["risultati"]["stats"].get("cronologia")
                if cronologia is not None and modifiche_log:
                    registra_rettifica_manuale(cronologia, modifiche_log)
                st.success(f"{n_mod} assegnazioni modificate. Riepiloghi ricalcolati.")
                st.rerun()
        else:
            st.dataframe(df_display, use_container_width=True, hide_index=True, height=500)

    with tab_riep:
        if not df_riep_display.empty:
            st.dataframe(df_riep_display, use_container_width=True, hide_index=True, height=400)
        else:
            st.info("Nessun dato nel riepilogo per il filtro selezionato.")

    with tab_crit:
        if not df_criticita.empty:
            st.dataframe(df_criticita, use_container_width=True, hide_index=True, height=400)
        else:
            st.success("Nessuna criticita rilevata.")

    with tab_graf:
        _mostra_grafici(df_assegnazioni, df_riepilogo,
                        n_settimane, perc_decurtazione, costo_orario, aliquota_iva)

    with tab_log:
        for entry in log_lines:
            st.text(entry)

    st.markdown("##### Documenti da scaricare")
    d1, d2, d3 = st.columns(3)
    with d1, st.container(border=True):
        st.markdown("**Report completo**")
        st.caption(
            "Un unico file Excel per l'ufficio: assegnazioni, riepilogo "
            "gruppi 45h, riepilogo economico, criticità e un foglio per "
            "ogni organismo."
        )
        excel_bytes = genera_excel(
            df_assegnazioni, df_riepilogo, df_criticita,
            n_settimane, perc_decurtazione, costo_orario, aliquota_iva,
        )
        st.download_button(
            "Scarica report completo (.xlsx)",
            data=excel_bytes,
            file_name="assegnazioni_output.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True,
            icon=":material/download:",
        )
    with d2, st.container(border=True):
        st.markdown("**File per organismo**")
        st.caption(
            "Un file Excel per ogni organismo, con i soli alunni di sua "
            "competenza: pronto per l'invio diretto nel rispetto della "
            "privacy (GDPR)."
        )
        zip_xlsx = genera_zip_excel_cooperative(
            df_assegnazioni, df_riepilogo, df_criticita,
            n_settimane, perc_decurtazione, costo_orario, aliquota_iva,
        )
        st.download_button(
            "Scarica file separati (.zip)",
            data=zip_xlsx,
            file_name="assegnazioni_per_cooperativa.zip",
            mime="application/zip",
            use_container_width=True,
            icon=":material/folder_zip:",
        )
    with d3, st.container(border=True):
        st.markdown("**Lettere di assegnazione**")
        st.caption(
            "Una lettera PDF per ogni organismo con il dettaglio degli "
            "alunni assegnati e il computo economico, da protocollare "
            "e trasmettere."
        )
        zip_pdf = genera_zip_pdf_cooperative(
            df_assegnazioni, n_settimane, perc_decurtazione, costo_orario, aliquota_iva,
        )
        st.download_button(
            "Scarica lettere (.zip PDF)",
            data=zip_pdf,
            file_name="lettere_assegnazione.zip",
            mime="application/zip",
            use_container_width=True,
            icon=":material/picture_as_pdf:",
        )

    cronologia = stats.get("cronologia")
    if cronologia and cronologia.get("passi"):
        meta_mesis = res.get("meta_mesis", {})
        parametri = dict(res.get("parametri", {}))
        parametri.setdefault("soglia", 45)
        parametri["corso_anno"] = res.get("corso_anno", parametri.get("corso_anno", False))
        n_passi = len(cronologia["passi"])
        n_mov = len(cronologia.get("movimenti", []))

        st.markdown("##### Trasparenza e accesso agli atti")
        st.caption(
            "Documentazione dell'intero procedimento, per ricostruire il "
            f"«prima e dopo» delle assegnazioni: {n_passi} fasi registrate, "
            f"{n_mov} spostamenti tracciati. Utile in caso di accesso agli atti "
            "(L. 241/1990) e a evidenza della logica del trattamento "
            "automatizzato (art. 22 Reg. UE 2016/679)."
        )

        # genera i documenti una sola volta per stato della cronologia: la
        # firma cambia solo su nuova esecuzione o rettifica manuale, così i
        # rerun non li rigenerano inutilmente
        firma_cron = (
            n_passi, n_mov,
            sum(len(p.get("stato", {})) for p in cronologia["passi"]),
            tuple(sorted(parametri.items(), key=lambda kv: kv[0])),
        )
        cache = st.session_state.get("_cron_docs_cache")
        if not cache or cache.get("firma") != firma_cron:
            cache = {
                "firma": firma_cron,
                "xlsx": genera_excel_cronologia(
                    cronologia, log_lines, meta_mesis, parametri),
                "pdf": genera_pdf_verbale(
                    cronologia, stats, log_lines, meta_mesis, parametri),
            }
            st.session_state["_cron_docs_cache"] = cache

        t1, t2 = st.columns(2)
        with t1, st.container(border=True):
            st.markdown("**Cronologia del procedimento**")
            st.caption(
                "Foglio Excel con la sintesi delle fasi, il diario per alunno "
                "(l'organismo ad ogni passaggio, con le variazioni evidenziate) "
                "e il registro puntuale di ogni spostamento con la motivazione."
            )
            st.download_button(
                "Scarica cronologia (.xlsx)",
                data=cache["xlsx"],
                file_name="cronologia_procedimento_oepac.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                icon=":material/history:",
            )
        with t2, st.container(border=True):
            st.markdown("**Verbale del procedimento**")
            st.caption(
                "Documento PDF formale che descrive fasi, spostamenti e "
                "riferimenti normativi del procedimento automatizzato, "
                "con spazio per data e firma del responsabile: pronto da "
                "allegare a una risposta di accesso agli atti."
            )
            st.download_button(
                "Scarica verbale (.pdf)",
                data=cache["pdf"],
                file_name="verbale_procedimento_oepac.pdf",
                mime="application/pdf",
                use_container_width=True,
                icon=":material/gavel:",
            )

    st.markdown("---")
    st.caption(
        "Gestionale OEPAC · Linee Guida DGC Roma Capitale n. 260/2024 · "
        "I dati trattati riguardano alunni con disabilità (art. 9 Reg. UE "
        "2016/679): conservare e trasmettere i file scaricati nel rispetto "
        "della normativa sulla protezione dei dati personali."
    )


def _mostra_grafici(df_ass, df_riep, n_settimane, perc_decurtazione, costo_orario, aliquota_iva):
    import altair as alt

    ORG = "Organismo Assegnato dall'Algoritmo"
    df_assegnati = df_ass[df_ass[ORG].astype(str).str.len() > 0]
    if df_assegnati.empty:
        st.info("Nessun dato da visualizzare.")
        return

    df_eco = calcola_colonne_economiche(
        df_assegnati, n_settimane, perc_decurtazione, costo_orario, aliquota_iva
    )

    g1, g2 = st.columns(2)
    with g1:
        st.markdown("**Ore settimanali per cooperativa**")
        per_org = df_eco.groupby(ORG)["Ore Assegnate"].sum().reset_index()
        per_org.columns = ["Cooperativa", "Ore"]
        chart = alt.Chart(per_org).mark_bar(color="#2E75B6").encode(
            x=alt.X("Ore:Q", title="Ore settimanali"),
            y=alt.Y("Cooperativa:N", sort="-x", title=None),
            tooltip=["Cooperativa", "Ore"],
        ).properties(height=max(150, 40 * len(per_org)))
        st.altair_chart(chart, use_container_width=True)

    with g2:
        st.markdown("**Alunni per cooperativa**")
        per_org_n = df_eco.groupby(ORG).size().reset_index(name="Alunni")
        per_org_n.columns = ["Cooperativa", "Alunni"]
        chart2 = alt.Chart(per_org_n).mark_bar(color="#1F4E79").encode(
            x=alt.X("Alunni:Q", title="N. alunni"),
            y=alt.Y("Cooperativa:N", sort="-x", title=None),
            tooltip=["Cooperativa", "Alunni"],
        ).properties(height=max(150, 40 * len(per_org_n)))
        st.altair_chart(chart2, use_container_width=True)

    g3, g4 = st.columns(2)
    with g3:
        st.markdown("**Importo totale annuo per cooperativa (EUR)**")
        per_org_eur = df_eco.groupby(ORG)["Totale (EUR)"].sum().reset_index()
        per_org_eur.columns = ["Cooperativa", "Totale"]
        chart3 = alt.Chart(per_org_eur).mark_bar(color="#2E9B5B").encode(
            x=alt.X("Totale:Q", title="EUR", axis=alt.Axis(format=",.0f")),
            y=alt.Y("Cooperativa:N", sort="-x", title=None),
            tooltip=["Cooperativa", alt.Tooltip("Totale:Q", format=",.2f")],
        ).properties(height=max(150, 40 * len(per_org_eur)))
        st.altair_chart(chart3, use_container_width=True)

    with g4:
        st.markdown("**Preferenze soddisfatte (nuove iscrizioni)**")
        pref = df_eco[df_eco["Preferenza Soddisfatta"].isin(
            ["1ª", "2ª", "3ª", "4ª", "5ª", "Manuale", "Non assegnato", "Già attivato"]
        )]["Preferenza Soddisfatta"].value_counts().reset_index()
        pref.columns = ["Preferenza", "Alunni"]
        if not pref.empty:
            chart4 = alt.Chart(pref).mark_arc(innerRadius=50).encode(
                theta="Alunni:Q",
                color=alt.Color("Preferenza:N", scale=alt.Scale(scheme="blues")),
                tooltip=["Preferenza", "Alunni"],
            ).properties(height=250)
            st.altair_chart(chart4, use_container_width=True)
        else:
            st.caption("Nessuna nuova iscrizione da mostrare.")

    st.markdown("**Distribuzione per grado scolastico**")
    grado = df_eco.groupby("Grado Scolastico").agg(
        Alunni=("Codice Iscrizione", "count"),
        Ore=("Ore Assegnate", "sum"),
    ).reset_index()
    st.dataframe(grado, use_container_width=True, hide_index=True)

    if "Tipo Gestione" in df_eco.columns:
        st.markdown("**Distribuzione per tipo gestione e municipio**")
        tipmun = df_eco.groupby(["Tipo Gestione", "Municipio"]).agg(
            Alunni=("Codice Iscrizione", "count"),
            Ore=("Ore Assegnate", "sum"),
            Gruppi=("Gruppo 45h", "nunique"),
        ).reset_index()
        st.dataframe(tipmun, use_container_width=True, hide_index=True)


if __name__ == "__main__":
    main()
