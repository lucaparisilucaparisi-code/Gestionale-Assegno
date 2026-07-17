"""Test del gestionale OEPAC: classificazione automatica ambiti (Roma),
modalità in corso d'anno e risoluzione per plesso.

Esecuzione:  python3 test_app.py

I test unitari e sintetici girano sempre. I test di integrazione sul file
MESIS reale (non incluso nel repo per privacy) girano solo se il file è
presente: esportarlo da MESIS e indicarne il percorso nella variabile
d'ambiente MESIS_XLSX (default: elenco_mesis.xlsx nella cartella corrente).
"""
import datetime
import io
import os
import sys

import pandas as pd

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import app  # noqa: E402

MESIS = os.environ.get("MESIS_XLSX", "elenco_mesis.xlsx")

FALLITI = []


def check(nome, cond, dettaglio=""):
    stato = "OK  " if cond else "FAIL"
    print(f"[{stato}] {nome}" + (f" — {dettaglio}" if dettaglio else ""))
    if not cond:
        FALLITI.append(nome)


# ---------------------------------------------------------------------------
# Unit test
# ---------------------------------------------------------------------------

def test_unitari():
    print("=== Unit: normalizza_municipio ===")
    casi_mun = [
        ("MUNICIPIO ROMA V ", "V"),
        ("MUNICIPIO ROMA XIV", "XIV"),
        ("Municipio 7", "VII"),
        ("1", "I"),
        ("01", "I"),
        ("15", "XV"),
        ("16", ""),          # non esiste
        ("XIII", "XIII"),
        ("municipio roma iii", "III"),
        ("", ""),
        (None, ""),
        (float("nan"), ""),
        (7, "VII"),
        ("5.0", "V"),
    ]
    for val, atteso in casi_mun:
        got = app.normalizza_municipio(val)
        check(f"normalizza_municipio({val!r}) = {atteso!r}", got == atteso, f"ottenuto {got!r}")

    print("\n=== Unit: estrai_numero_ambito ===")
    for val, atteso in [("10", "10"), ("12 - Paritario", "12"), ("10.0", "10"),
                        ("IC", ""), ("", ""), (None, "")]:
        got = app.estrai_numero_ambito(val)
        check(f"estrai_numero_ambito({val!r}) = {atteso!r}", got == atteso, f"ottenuto {got!r}")

    print("\n=== Unit: classifica_gestione ===")
    casi_class = [
        ("RMEE8E101X", "10", "VIA ARETUSA", "LA RUSTICA", None, "Statale", "codice"),
        ("RMAA82301D", "IC", "I.C. VALENTE", "MATERNA", None, "Statale", "codice"),
        ("RMMM8EW01X", "IC", "IC X", "Y", None, "Statale", "codice"),
        ("RM1A00900R", "11", "F. CECCONI - COMUNALE", "F. CECCONI - COMUNALE", None,
         "Infanzia comunale", "testo"),
        ("RM1A23100Q", "12", "ROMITI", "ROMITI", None, "Infanzia comunale", "ambito MESIS"),
        ("RM1A26500V", "12 - Paritario", "SC. INF. PARIT S.P. CERIOLI", "idem", None,
         "Paritaria", "testo"),
        ("RM1A304008", "10 - Paritario", "SCUOLA MATERNA NOSTRA SIGNORA", "idem", None,
         "Paritaria", "testo"),
        ("RM1E00200X", "11 - Paritario", "SCUOLA ELEM.PARIT. PIO XII", "idem", None,
         "Paritaria", "codice"),
        ("RM1M08300T", "11", "SCUOLA MEDIA CABRINI", "idem", None, "Paritaria", "codice"),
        ("RM1A99999X", "", "SCONOSCIUTA", "SCONOSCIUTA", None, "Infanzia comunale", "presunta"),
        ("RM1A99999X", "", "SCONOSCIUTA", "SCONOSCIUTA", {"RM1A99999X"},
         "Infanzia comunale", "elenco comunali"),
        ("", "", "", "", None, "N/D", ""),
        ("XX", "", "", "", None, "N/D", ""),
    ]
    for cod, amb, ist, ple, elenco, t_att, f_att in casi_class:
        t, f = app.classifica_gestione(cod, amb, ist, ple, elenco)
        check(
            f"classifica_gestione({cod!r}, {amb!r}) = ({t_att!r}, {f_att!r})",
            (t, f) == (t_att, f_att), f"ottenuto ({t!r}, {f!r})",
        )

    print("\n=== Unit: deriva_gruppo_auto ===")
    casi_grp = [
        (("Statale", "V", "RMIC8E0001", "RMEE8E101X", "10"), "IC:RMIC8E0001"),
        (("Statale", "V", "", "RMEE8E101X", "10"), "PLESSO:RMEE8E101X"),
        (("Infanzia comunale", "V", "RM1A23100Q", "RM1A23100Q", "12"), "COM:Ambito 12"),
        (("Infanzia comunale", "", "RM1A23100Q", "RM1A23100Q", "12"), "COM:Ambito 12"),
        (("Infanzia comunale", "V", "X", "X", ""), "COM:Municipio V"),
        (("Infanzia comunale", "", "X", "X", ""), "COM:Municipio N/D"),
        (("Paritaria", "V", "X", "RM1A26500V", "12 - Paritario"), "PAR:Ambito 12"),
        (("Paritaria", "V", "X", "RM1A26500V", ""), "PAR:Municipio V"),
        (("N/D", "V", "X", "PLX", "qualcosa"), "AMB:qualcosa"),
        (("N/D", "V", "X", "PLX", ""), "PLESSO:PLX"),
    ]
    for args, atteso in casi_grp:
        got = app.deriva_gruppo_auto(*args)
        check(f"deriva_gruppo_auto{args} = {atteso!r}", got == atteso, f"ottenuto {got!r}")

    print("\n=== Unit: deriva_grado ===")
    for cod, ctx, atteso in [
        ("RMAA8E001T", "Statale", "Infanzia statale"),
        ("RMEE8E101X", "Statale", "Primaria statale"),
        ("RMMM8EW01X", "Statale", "Sec. I grado statale"),
        ("RM1A23100Q", "Infanzia comunale", "Infanzia comunale"),
        ("RM1A26500V", "Paritaria", "Infanzia paritaria"),
        ("RM1E00200X", "Paritaria", "Primaria paritaria"),
        ("RM1M08300T", "Paritaria", "Sec. I grado paritaria"),
    ]:
        got = app.deriva_grado(cod, ctx)
        check(f"deriva_grado({cod!r}, {ctx!r}) = {atteso!r}", got == atteso, f"ottenuto {got!r}")


# ---------------------------------------------------------------------------
# Integrazione su un file MESIS reale (opzionale)
# ---------------------------------------------------------------------------

def test_file_reale(percorso):
    print("\n=== Integrazione: file MESIS reale ===")
    with open(percorso, "rb") as f:
        file_bytes = f.read()

    df_raw, col_map, errors = app.carica_dati(file_bytes)
    check("carica_dati senza errori", not errors, str(errors))

    res = app.esegui_assegnazione(df_raw, col_map, 45, 50, auto_ambito=True)
    df_ass, df_riep, df_crit, log, stats, registro = res

    statali = df_ass[df_ass["Tipo Gestione"] == "Statale"]
    comunali = df_ass[df_ass["Tipo Gestione"] == "Infanzia comunale"]
    paritarie = df_ass[df_ass["Tipo Gestione"] == "Paritaria"]
    nd = df_ass[df_ass["Tipo Gestione"] == "N/D"]
    print(f"    statali={len(statali)} comunali={len(comunali)} "
          f"paritarie={len(paritarie)} N/D={len(nd)}")
    check("nessun N/D", len(nd) == 0)
    check("statali tutti in gruppi IC:", statali["Gruppo 45h"].str.startswith("IC:").all())
    check("comunali tutti in gruppi COM:", comunali["Gruppo 45h"].str.startswith("COM:").all())
    check("paritarie tutte in gruppi PAR:", paritarie["Gruppo 45h"].str.startswith("PAR:").all())
    n_ic = statali["Gruppo 45h"].nunique()
    n_ic_codici = statali["Codice Meccanografico Istituto"].nunique()
    check("gruppi IC = istituti distinti", n_ic == n_ic_codici, f"{n_ic} vs {n_ic_codici}")

    check("riepilogo ha colonna Municipio", "Municipio" in df_riep.columns)
    tipi_grp = set(df_riep["Tipo gruppo"].unique())
    check("riepilogo con tipi IC/Infanzia comunale/Paritario",
          {"IC", "Infanzia comunale", "Paritario"} <= tipi_grp, str(tipi_grp))

    ore_in = pd.to_numeric(df_raw[col_map["ore_assegnate"]], errors="coerce").fillna(0)
    mask_att = df_raw[col_map["stato"]].str.strip().str.upper() == "ATTIVA"
    rin = col_map.get("data_rinuncia")
    if rin:
        mask_att &= df_raw[rin].isna() | (df_raw[rin].str.strip() == "")
    check("somma ore input == output",
          abs(ore_in[mask_att].sum() - df_ass["Ore Assegnate"].sum()) < 0.01)

    checks = app.verifiche_consistenza(df_ass, df_raw, col_map, stats, auto_ambito=True)
    for label, ok, det in checks:
        print(f"    [{'ok' if ok else 'ATTN'}] {label}: {det}")

    xlsx = app.genera_excel(df_ass, df_riep, df_crit)
    check("genera_excel gira", len(xlsx) > 10000, f"{len(xlsx)} bytes")

    df_scuole = app.costruisci_tabella_scuole(df_raw, col_map, None)
    check("tabella scuole non vuota", len(df_scuole) > 0, str(len(df_scuole)))

    # corso d'anno: gli attivati restano sul loro organismo
    res_ca = app.esegui_assegnazione(df_raw, col_map, 45, 50,
                                     auto_ambito=True, corso_anno=True)
    df_ca, _, _, log_ca, stats_ca, _ = res_ca
    att = df_ca[df_ca["Preferenza Soddisfatta"] == "Già attivato"]
    check("corso d'anno: attivati mantenuti sul loro organismo",
          (att["Organismo Assegnato dall'Algoritmo"] == att["Organismo Pre-esistente"]).all())
    check("corso d'anno: conteggio attivati coerente",
          stats_ca["gia_attivati"] == len(att), f"{stats_ca['gia_attivati']} vs {len(att)}")

    # retrocompatibilità: gruppi legacy
    res_leg = app.esegui_assegnazione(df_raw, col_map, 45, 50, auto_ambito=False)
    grp_leg = set(res_leg[0]["Gruppo 45h"].unique())
    check("legacy: soli prefissi IC:/AMB:/PLESSO:",
          all(g.startswith(("IC:", "AMB:", "PLESSO:")) for g in grp_leg))


# ---------------------------------------------------------------------------
# Sintetici: ambito/municipio mancanti, multi-municipio, risoluzione per plesso
# ---------------------------------------------------------------------------

INTESTAZIONI = [
    "Codice", "Stato", "Tipo", "Municipio", "Cognome", "Nome",
    "Data Nascita", "Codice Meccanografico Istituto", "Istituto",
    "Codice Meccanografico Plesso", "Plesso", "Classe", "Sezione",
    "Ambito", "Ore Richieste", "Ore Assegnate", "UTENTE",
    "Organismo Assegnato", "Enti Erogatori scelti", "Data Attivazione",
    "Data Rinuncia/Sospensione",
]


def costruisci_mesis(rows):
    """Costruisce un finto xlsx MESIS (7 righe metadati + header + dati)."""
    meta = [["x"] + [None] * (len(INTESTAZIONI) - 1)] * 7
    data = list(meta) + [INTESTAZIONI] + [
        [r.get(h) for h in INTESTAZIONI] for r in rows
    ]
    buf = io.BytesIO()
    pd.DataFrame(data).to_excel(buf, index=False, header=False)
    return buf.getvalue()


def riga(codice, tipo, mun, cod_ist, ist, cod_ple, ple, ambito, ore,
         org="", enti="", att=""):
    return {
        "Codice": codice, "Stato": "ATTIVA", "Tipo": tipo, "Municipio": mun,
        "Cognome": f"C{codice}", "Nome": f"N{codice}", "Data Nascita": "01/01/2018",
        "Codice Meccanografico Istituto": cod_ist, "Istituto": ist,
        "Codice Meccanografico Plesso": cod_ple, "Plesso": ple,
        "Classe": "1", "Sezione": "A", "Ambito": ambito,
        "Ore Richieste": str(ore), "Ore Assegnate": str(ore),
        "UTENTE": f"U{codice}",
        "Organismo Assegnato": org, "Enti Erogatori scelti": enti,
        "Data Attivazione": att, "Data Rinuncia/Sospensione": None,
    }


def test_sintetici():
    print("\n=== Sintetico: ambito mancante e multi-municipio ===")
    rows = [
        riga("1", "RICONFERMA", "MUNICIPIO ROMA VII", "RMIC000001", "IC ALFA",
             "RMEE000011", "PLESSO ALFA", "", 20, org="COOP A"),
        riga("2", "RICONFERMA", "MUNICIPIO ROMA VII", "RM1A000002", "ARCOBALENO",
             "RM1A000002", "ARCOBALENO", "", 22, org="COOP B"),
        riga("3", "RICONFERMA", "MUNICIPIO ROMA VII", "RM1A000003", "SC. INF. PARITARIA BETA",
             "RM1A000003", "SC. INF. PARITARIA BETA", "", 15, org="COOP C"),
        riga("4", "RICONFERMA", "MUNICIPIO ROMA III", "RM1A000004", "GAMMA - COMUNALE",
             "RM1A000004", "GAMMA - COMUNALE", "5", 18, org="COOP D"),
        riga("5", "RICONFERMA", "MUNICIPIO ROMA VII", "RM1A000005", "DELTA - COMUNALE",
             "RM1A000005", "DELTA - COMUNALE", "5", 19, org="COOP E"),
        riga("6", "RICONFERMA", "", "RM1A000006", "EPSILON - COMUNALE",
             "RM1A000006", "EPSILON - COMUNALE", "", 12, org="COOP F"),
    ]
    df_s, cm_s, err_s = app.carica_dati(costruisci_mesis(rows))
    check("sintetico: carica_dati ok", not err_s, str(err_s))
    res_s = app.esegui_assegnazione(df_s, cm_s, 45, 50, auto_ambito=True)
    df_rs = res_s[0]
    by_code = df_rs.set_index("Codice Iscrizione")
    casi = [
        ("1", "Statale", "IC:RMIC000001"),
        ("2", "Infanzia comunale", "COM:Municipio VII"),
        ("3", "Paritaria", "PAR:Municipio VII"),
        ("4", "Infanzia comunale", "COM:Ambito 5"),
        ("5", "Infanzia comunale", "COM:Ambito 5"),
        ("6", "Infanzia comunale", "COM:Municipio N/D"),
    ]
    for cod, tipo_att, grp_att in casi:
        t = by_code.at[cod, "Tipo Gestione"]
        g = by_code.at[cod, "Gruppo 45h"]
        check(f"sintetico riga {cod}: {tipo_att} / {grp_att}",
              (t, g) == (tipo_att, grp_att), f"ottenuto ({t!r}, {g!r})")
    check("sintetico riga 2 fonte presunta",
          by_code.at["2", "Fonte Classificazione"] == "presunta")

    # ambito 5 presente in due municipi (III e VII): il check deve avvisare
    checks_s = app.verifiche_consistenza(df_rs, df_s, cm_s, res_s[4], auto_ambito=True)
    chk_amb = [c for c in checks_s if c[0] == "Ambiti territoriali univoci tra municipi"]
    check("check ambiti multi-municipio in allerta",
          len(chk_amb) == 1 and chk_amb[0][1] is False, str(chk_amb))


def test_vincolo_45h_riassegnazione():
    print("\n=== Vincolo 45h in riassegnazione (riconferme contano ma non esentano) ===")
    G = ("MUNICIPIO ROMA V", "RMIC0000IC", "IC TEST", "RMEE0000IC", "PLESSO", "IC")

    def r_ric(cod, ore, org):
        return riga(cod, "RICONFERMA", *G, ore, org=org)

    def r_nuova(cod, ore, prefs):
        return riga(cod, "NUOVA ISCRIZIONE", *G, ore, enti="\n".join(prefs))

    ORG = "Organismo Assegnato dall'Algoritmo"

    def esegui(rows):
        df, cm, _ = app.carica_dati(costruisci_mesis(rows))
        return app.esegui_assegnazione(df, cm, 45, 50, auto_ambito=True)[0].set_index("Codice Iscrizione")

    # Scenario 1: una coop con SOLA riconferma sotto 45h non è valida per una
    # nuova; la nuova va scalata alla preferenza successiva che raggiunge 45h.
    r = esegui([
        r_ric("R1", 30, "Coop-X"),   # Coop-X: 30h di riconferma
        r_ric("R2", 50, "Coop-Z"),   # Coop-Z: 50h di riconferma (valida)
        r_nuova("A", 10, ["Coop-X", "Coop-Z"]),
    ])
    check("scenario 1: nuova NON resta su coop sotto 45h con sola riconferma",
          r.at["A", ORG] == "Coop-Z" and r.at["A", "Preferenza Soddisfatta"] == "2ª",
          f"A -> {r.at['A', ORG]} ({r.at['A', 'Preferenza Soddisfatta']})")

    # Scenario 2: la nuova si aggiunge a una coop con riconferma portandola a
    # 45h (le ore delle riconferme concorrono alla soglia).
    r2 = esegui([
        r_ric("R1", 40, "Coop-X"),
        r_ric("R2", 60, "Coop-W"),
        r_nuova("A", 10, ["Coop-Y", "Coop-X"]),  # Coop-Y solo A (10h), poi Coop-X
    ])
    check("scenario 2: nuova si aggiunge alla riconferma (40+10=50h, valida)",
          r2.at["A", ORG] == "Coop-X" and r2.at["A", "Status"] == "OK",
          f"A -> {r2.at['A', ORG]} status={r2.at['A', 'Status']}")

    # Scenario 3: nessuna preferenza raggiunge 45h -> da assegnare manualmente
    r3 = esegui([
        r_ric("R1", 60, "Coop-BIG"),  # tiene il gruppo sopra 45h totali
        r_nuova("A", 10, ["Coop-P", "Coop-Q"]),  # entrambe solo A (10h)
    ])
    check("scenario 3: nessuna preferenza a 45h -> da assegnare manualmente",
          r3.at["A", "Status"] == "Da assegnare manualmente",
          f"A status={r3.at['A', 'Status']}")

    # Scenario 4: gruppo con monte ore totale sotto 45h -> eccezione, nessuno spostato
    r4 = esegui([
        r_nuova("A", 10, ["Coop-P", "Coop-Q"]),
        r_nuova("B", 15, ["Coop-P"]),
    ])
    check("scenario 4: gruppo sotto 45h totali -> nuove restano sulla 1ª (eccezione)",
          r4.at["A", "Preferenza Soddisfatta"] == "1ª"
          and r4.at["B", "Preferenza Soddisfatta"] == "1ª",
          f"A={r4.at['A', 'Preferenza Soddisfatta']} B={r4.at['B', 'Preferenza Soddisfatta']}")

    # Proprietà generale: nessuna nuova assegnata resta su una coop sotto 45h
    # (quando il gruppo raggiunge 45h totali). Riconferme escluse (continuità).
    import re as _re

    def _norm(s):
        return _re.sub(r"\s+", " ", str(s).strip()).casefold()

    for scen, rows in [("multi", [
        r_ric("R1", 30, "Coop-X"), r_ric("R2", 20, "Coop-Y"),
        r_nuova("A", 10, ["Coop-X", "Coop-Y", "Coop-Z"]),
        r_nuova("B", 40, ["Coop-Z"]),
        r_nuova("C", 10, ["Coop-Y", "Coop-Z"]),
    ])]:
        rr = esegui(rows)
        gtot = rr["Ore Assegnate"].sum()
        ore = {}
        for cod, row in rr.iterrows():
            if row[ORG]:
                ore[_norm(row[ORG])] = ore.get(_norm(row[ORG]), 0) + row["Ore Assegnate"]
        viol = 0
        for cod, row in rr.iterrows():
            if not row[ORG] or row["Preferenza Soddisfatta"] in ("Riconferma", "Già attivato"):
                continue
            if gtot >= 45 and ore[_norm(row[ORG])] < 45:
                viol += 1
        check(f"proprietà ({scen}): nessuna nuova assegnata sotto 45h", viol == 0, f"{viol} violazioni")


def test_risoluzione_per_plesso():
    print("\n=== Risoluzione per plesso (regressioni review) ===")

    # municipio mancante su UNA riga: gruppo unico e criticità preservata
    rows_f1 = [
        riga("11", "RICONFERMA", "MUNICIPIO ROMA V", "RM1A000011", "ALFA - COMUNALE",
             "RM1A000011", "ALFA - COMUNALE", "10", 20, org="COOP UNO"),
        riga("12", "RICONFERMA", "MUNICIPIO ROMA V", "RM1A000011", "ALFA - COMUNALE",
             "RM1A000011", "ALFA - COMUNALE", "10", 20, org="COOP UNO"),
        riga("13", "RICONFERMA", "MUNICIPIO ROMA V", "RM1A000011", "ALFA - COMUNALE",
             "RM1A000011", "ALFA - COMUNALE", "10", 20, org="COOP UNO"),
        riga("14", "RICONFERMA", "", "RM1A000011", "ALFA - COMUNALE",
             "RM1A000011", "ALFA - COMUNALE", "10", 20, org="COOP DUE"),
    ]
    df_f1, cm_f1, _ = app.carica_dati(costruisci_mesis(rows_f1))
    df_rf1 = app.esegui_assegnazione(df_f1, cm_f1, 45, 50, auto_ambito=True)[0]
    check("municipio mancante su una riga: un solo gruppo COM:Ambito 10",
          set(df_rf1["Gruppo 45h"].unique()) == {"COM:Ambito 10"},
          str(set(df_rf1["Gruppo 45h"].unique())))
    check("criticità sotto soglia rilevata (20h < 45h su gruppo da 80h)",
          (df_rf1["Status"] == "Riconferma sotto soglia").sum() == 1,
          str(df_rf1["Status"].value_counts().to_dict()))
    check("municipio risolto anche sulla riga senza valore",
          (df_rf1["Municipio"] == "V").all())

    # ambito incoerente tra righe dello stesso plesso: classificazione unica
    rows_f2 = [
        riga("21", "RICONFERMA", "MUNICIPIO ROMA V", "RM1A000021", "BETA",
             "RM1A000021", "BETA", "12 - Paritario", 20, org="COOP TRE"),
        riga("22", "RICONFERMA", "MUNICIPIO ROMA V", "RM1A000021", "BETA",
             "RM1A000021", "BETA", "12", 20, org="COOP TRE"),
    ]
    df_f2, cm_f2, _ = app.carica_dati(costruisci_mesis(rows_f2))
    df_rf2 = app.esegui_assegnazione(df_f2, cm_f2, 45, 50, auto_ambito=True)[0]
    check("ambito incoerente: tipo unico Paritaria",
          set(df_rf2["Tipo Gestione"].unique()) == {"Paritaria"},
          str(set(df_rf2["Tipo Gestione"].unique())))
    check("ambito incoerente: gruppo unico PAR:Ambito 12",
          set(df_rf2["Gruppo 45h"].unique()) == {"PAR:Ambito 12"},
          str(set(df_rf2["Gruppo 45h"].unique())))
    tab_f2 = app.costruisci_tabella_scuole(df_f2, cm_f2, None)
    check("tabella scuole allineata al motore",
          len(tab_f2) == 1
          and tab_f2.iloc[0]["Tipo Gestione"] == "Paritaria"
          and tab_f2.iloc[0]["Gruppo 45h (auto)"] == "PAR:Ambito 12")

    # override manuale per plesso
    ovr = app.esegui_assegnazione(
        df_f2, cm_f2, 45, 50, auto_ambito=True,
        tipi_override={"RM1A000021": "Infanzia comunale"},
    )[0]
    check("override manuale applicato a tutte le righe del plesso",
          set(ovr["Tipo Gestione"].unique()) == {"Infanzia comunale"}
          and set(ovr["Gruppo 45h"].unique()) == {"COM:Ambito 12"}
          and set(ovr["Fonte Classificazione"].unique()) == {"manuale"})


def test_corso_anno_sintetico():
    print("\n=== Sintetico: corso d'anno ===")
    rows = [
        # riconferma
        riga("31", "RICONFERMA", "MUNICIPIO ROMA V", "RMIC000031", "IC GAMMA",
             "RMEE000031", "PLESSO GAMMA", "IC", 20, org="COOP A",
             att="2025-09-15 00:00:00"),
        # nuova già attivata a novembre (finestra 2)
        riga("32", "NUOVA ISCRIZIONE", "MUNICIPIO ROMA V", "RMIC000031", "IC GAMMA",
             "RMEE000031", "PLESSO GAMMA", "IC", 25, org="COOP A",
             enti="COOP B\nCOOP A", att="2025-11-17 00:00:00"),
        # nuova della finestra corrente, da assegnare
        riga("33", "NUOVA ISCRIZIONE", "MUNICIPIO ROMA V", "RMIC000031", "IC GAMMA",
             "RMEE000031", "PLESSO GAMMA", "IC", 10, enti="COOP A\nCOOP B"),
    ]
    df_c, cm_c, _ = app.carica_dati(costruisci_mesis(rows))
    res_c = app.esegui_assegnazione(df_c, cm_c, 45, 50,
                                    auto_ambito=True, corso_anno=True)
    df_rc, stats_c = res_c[0], res_c[4]
    by = df_rc.set_index("Codice Iscrizione")
    check("attivato (32) mantenuto su COOP A nonostante 1ª preferenza COOP B",
          by.at["32", "Organismo Assegnato dall'Algoritmo"] == "COOP A"
          and by.at["32", "Preferenza Soddisfatta"] == "Già attivato")
    check("nuovo (33) assegnato dalla 1ª preferenza",
          by.at["33", "Organismo Assegnato dall'Algoritmo"] == "COOP A")
    check("conteggio attivati = 1", stats_c["gia_attivati"] == 1)

    # con data di riferimento al 30/10: l'attivazione di novembre non è fissa
    res_d = app.esegui_assegnazione(
        df_c, cm_c, 45, 50, auto_ambito=True, corso_anno=True,
        data_riferimento=datetime.date(2025, 10, 30),
    )
    check("data riferimento esclude l'attivazione di novembre",
          res_d[4]["gia_attivati"] == 0, str(res_d[4]["gia_attivati"]))


def test_cronologia():
    print("\n=== Cronologia del procedimento (trasparenza) ===")

    # scenario con uno spostamento forzato dal vincolo 45h:
    # nel gruppo IC, PICCOLA (10h, 1ª pref) sotto soglia -> spostata a GRANDE
    rows = [
        # riconferma che porta il gruppo sopra 45h totali
        riga("41", "RICONFERMA", "MUNICIPIO ROMA V", "RMIC000041", "IC DELTA",
             "RMEE000041", "PLESSO DELTA", "IC", 40, org="GRANDE Coop"),
        # nuova con 1ª pref PICCOLA (che resterà sotto 45h) e 2ª GRANDE
        riga("42", "NUOVA ISCRIZIONE", "MUNICIPIO ROMA V", "RMIC000041", "IC DELTA",
             "RMEE000041", "PLESSO DELTA", "IC", 10,
             enti="PICCOLA Coop\nGRANDE Coop"),
        # altra nuova su GRANDE per superare la soglia
        riga("43", "NUOVA ISCRIZIONE", "MUNICIPIO ROMA V", "RMIC000041", "IC DELTA",
             "RMEE000041", "PLESSO DELTA", "IC", 20, enti="GRANDE Coop"),
    ]
    df_x, cm_x, _ = app.carica_dati(costruisci_mesis(rows))
    res_x = app.esegui_assegnazione(df_x, cm_x, 45, 50, auto_ambito=True)
    stats_x = res_x[4]
    cron = stats_x.get("cronologia")
    check("cronologia presente in stats", cron is not None)
    passi = cron["passi"]
    titoli = [p["titolo"] for p in passi]
    check("prima fase = stato iniziale", titoli[0].startswith("Stato iniziale"))
    check("seconda fase = assegnazione iniziale", titoli[1] == "Assegnazione iniziale")
    check("ultima fase = esito definitivo", titoli[-1] == "Esito definitivo")
    # niente fasi/iterazioni intermedie: solo 3 fasi significative
    check("esattamente 3 fasi (niente tappe intermedie)", len(passi) == 3, str(titoli))
    check("nessuna fase 'iterazione' nella cronologia",
          not any("iterazione" in t.lower() for t in titoli), str(titoli))

    # l'alunno 42 nella cronologia passa da PICCOLA a GRANDE
    anagr = cron["anagrafica"]
    idx42 = [i for i, a in anagr.items() if a["codice"] == "42"][0]
    iniziale = passi[1]["stato"][idx42]
    finale = passi[-1]["stato"][idx42]
    check("42: assegnazione iniziale = PICCOLA Coop", iniziale == "PICCOLA Coop", iniziale)
    check("42: esito finale = GRANDE Coop", finale == "GRANDE Coop", finale)

    movimenti = cron["movimenti"]
    mov42 = [m for m in movimenti if anagr[m["idx"]]["codice"] == "42"]
    check("42: uno spostamento registrato con motivo", len(mov42) == 1
          and "soglia" in mov42[0]["motivo"].lower(), str(mov42))
    check("42: spostamento da PICCOLA a GRANDE",
          mov42 and mov42[0]["da"] == "PICCOLA Coop" and mov42[0]["a"] == "GRANDE Coop")

    # DataFrame della cronologia
    df_passi, df_diario, df_mov, intest = app.costruisci_cronologia_dfs(cron)
    check("df_diario ha una riga per alunno (3)", len(df_diario) == 3, str(len(df_diario)))
    check("df_diario ha una colonna per fase", all(h in df_diario.columns for h in intest))
    check("df_movimenti non vuoto", len(df_mov) >= 1)
    riga42 = df_diario[df_diario["Codice Iscrizione"] == "42"].iloc[0]
    check("diario 42: fase iniziale mostra PICCOLA",
          riga42[intest[1]] == "PICCOLA Coop", riga42[intest[1]])
    check("diario 42: fase finale mostra GRANDE",
          riga42[intest[-1]] == "GRANDE Coop", riga42[intest[-1]])
    check("diario 42: N. spostamenti = 1", int(riga42["N. spostamenti"]) == 1,
          str(riga42["N. spostamenti"]))

    # i conteggi devono riconciliare: fasi = diario = movimenti = algoritmo
    somma_fasi = sum(int(v) for v in df_passi["Spostamenti in questa fase"]
                     if str(v).strip() != "")
    somma_diario = int(df_diario["N. spostamenti"].sum())
    check("conteggi riconciliano (fasi=diario=movimenti=stats)",
          somma_fasi == somma_diario == len(df_mov) == stats_x["spostamenti"] == 1,
          f"fasi={somma_fasi} diario={somma_diario} mov={len(df_mov)} stats={stats_x['spostamenti']}")
    # nessuno spostamento fantasma nella fase "Assegnazione iniziale"
    fase_iniz = df_passi[df_passi["Fase"] == "Assegnazione iniziale"].iloc[0]
    check("fase 'Assegnazione iniziale': 0 spostamenti fantasma",
          int(fase_iniz["Spostamenti in questa fase"]) == 0)
    # stato iniziale: una nuova iscrizione NON ha organismo (vuoto)
    idx43 = [i for i, a in anagr.items() if a["codice"] == "43"][0]
    check("stato iniziale: nuova iscrizione senza organismo",
          cron["passi"][0]["stato"][idx43] == "", repr(cron["passi"][0]["stato"][idx43]))
    # stato iniziale: la riconferma mostra l'organismo pre-esistente
    idx41 = [i for i, a in anagr.items() if a["codice"] == "41"][0]
    check("stato iniziale: riconferma mostra organismo pre-esistente",
          cron["passi"][0]["stato"][idx41] == "GRANDE Coop")

    # caratteri di controllo nei dati non devono far crashare l'Excel
    import copy
    cron_cc = copy.deepcopy(cron)
    prima_chiave = next(iter(cron_cc["anagrafica"]))
    cron_cc["anagrafica"][prima_chiave]["cognome"] = "ROSSI\x07\x00\x1f"
    xlsx_cc = app.genera_excel_cronologia(
        cron_cc, ["log\x00riga"], {"municipio": "X\x1f"}, {"soglia": 45})
    check("Excel generato con caratteri di controllo (nessun crash)", len(xlsx_cc) > 5000)

    # generazione documenti
    meta = {"municipio": "MUNICIPIO ROMA V", "anno": "2025/2026"}
    parametri = {"soglia": 45, "corso_anno": False,
                 "iterazioni": stats_x.get("iterazioni", 0),
                 "spostamenti": stats_x.get("spostamenti", 0)}
    xlsx = app.genera_excel_cronologia(cron, res_x[3], meta, parametri)
    check("genera_excel_cronologia produce un file", len(xlsx) > 5000, f"{len(xlsx)} bytes")
    pdf = app.genera_pdf_verbale(cron, stats_x, res_x[3], meta, parametri)
    check("genera_pdf_verbale produce un PDF",
          len(pdf) > 2000 and pdf[:4] == b"%PDF", f"{len(pdf)} bytes")

    # caso senza spostamenti: verbale/excel devono comunque generarsi
    rows_ok = [
        riga("51", "NUOVA ISCRIZIONE", "MUNICIPIO ROMA V", "RMIC000051", "IC EPS",
             "RMEE000051", "PLESSO EPS", "IC", 50, enti="UNICA Coop"),
    ]
    df_o, cm_o, _ = app.carica_dati(costruisci_mesis(rows_ok))
    res_o = app.esegui_assegnazione(df_o, cm_o, 45, 50, auto_ambito=True)
    cron_o = res_o[4]["cronologia"]
    _, _, df_mov_o, _ = app.costruisci_cronologia_dfs(cron_o)
    check("senza spostamenti: nessun movimento", len(df_mov_o) == 0)
    pdf_o = app.genera_pdf_verbale(cron_o, res_o[4], res_o[3], meta, parametri)
    check("verbale generato anche senza spostamenti", pdf_o[:4] == b"%PDF")

    # rettifica manuale documentata nella cronologia
    app.registra_rettifica_manuale(cron, [{
        "codice": "43", "cognome": "C43", "nome": "N43", "plesso": "PLESSO DELTA",
        "gruppo_label": "Istituto Comprensivo (RMIC000041)",
        "da": "GRANDE Coop", "a": "PICCOLA Coop",
    }])
    check("rettifica manuale aggiunge una fase",
          cron["passi"][-1]["titolo"] == "Rettifica manuale dell'ufficio")
    mov_man = [m for m in cron["movimenti"] if m["passo_titolo"] == "Rettifica manuale dell'ufficio"]
    check("rettifica manuale registra il movimento", len(mov_man) == 1)
    # dopo la rettifica i conteggi restano coerenti: 1 algoritmo + 1 manuale
    par_dopo = app._conteggi_movimenti(cron, {"soglia": 45})
    check("conteggi dopo rettifica: 1 algoritmo, 1 manuale",
          par_dopo["spostamenti"] == 1 and par_dopo["rettifiche_manuali"] == 1,
          str(par_dopo))
    pdf_dopo = app.genera_pdf_verbale(cron, stats_x, res_x[3], meta, {"soglia": 45})
    check("verbale rigenerato dopo rettifica manuale", pdf_dopo[:4] == b"%PDF")


def test_cronologia_movimento_netto():
    print("\n=== Cronologia: solo movimento netto (niente coop fantasma) ===")
    G = ("MUNICIPIO ROMA V", "RMIC0000IC", "IC TEST", "RMEE0000IC", "PLESSO", "IC")
    ORG = "Organismo Assegnato dall'Algoritmo"
    # A sceglie [P1, P2, P3]: P1 e P2 restano sotto 45h (solo A), P3 e' valida
    # (riconferma da 50h). A dovrebbe rimbalzare P1->P2->P3 internamente, ma
    # la cronologia deve mostrare solo da P1 a P3, senza P2 (fantasma).
    rows = [
        riga("R", "RICONFERMA", *G, 50, org="P3 Coop"),
        riga("A", "NUOVA ISCRIZIONE", *G, 10, enti="P1 Coop\nP2 Coop\nP3 Coop"),
    ]
    df, cm, _ = app.carica_dati(costruisci_mesis(rows))
    res = app.esegui_assegnazione(df, cm, 45, 50, auto_ambito=True)
    dfa, cron = res[0], res[4]["cronologia"]
    anagr = cron["anagrafica"]
    fin = dfa.loc[dfa["Codice Iscrizione"] == "A", ORG].iloc[0]
    check("A finisce su P3 Coop (3ª preferenza)", fin == "P3 Coop", fin)
    mov = [m for m in cron["movimenti"] if anagr[m["idx"]]["codice"] == "A"]
    check("A: un solo movimento netto registrato", len(mov) == 1, str(mov))
    check("A: movimento da P1 a P3 (niente P2 intermedia)",
          mov and mov[0]["da"] == "P1 Coop" and mov[0]["a"] == "P3 Coop",
          str(mov))
    dest = {m["a"] for m in cron["movimenti"] if m["a"]}
    check("nessuna coop intermedia 'P2 Coop' nei movimenti", "P2 Coop" not in dest)
    orgs_finali = {o for o in dfa[ORG].unique() if o}
    check("ogni destinazione dei movimenti e' un organismo finale",
          all(m["a"] in orgs_finali for m in cron["movimenti"] if m["a"]))


def test_attribuzione_report():
    print("\n=== Attribuzione dei fogli per-organismo nel report ===")
    from openpyxl import load_workbook
    ORG = "Organismo Assegnato dall'Algoritmo"

    def _riga_report(cod, org, ore):
        return {
            "Codice Iscrizione": cod, "Tipo": "NUOVA ISCRIZIONE",
            "Cognome": "C" + cod, "Nome": "N" + cod, "Data Nascita": "",
            "Età": 6, "Codice Meccanografico Istituto": "RMIC1", "Istituto": "IC TEST",
            "Codice Meccanografico Plesso": "RMEE1", "Plesso": "PLESSO",
            "Indirizzo Plesso": "", "Grado Scolastico": "Primaria statale",
            "Classe": "1", "Sezione": "A", "Ambito": "IC", "Tipo Gestione": "Statale",
            "Fonte Classificazione": "codice", "Municipio": "V", "Gruppo 45h": "IC:RMIC1",
            "Ore Richieste": ore, "Ore Assegnate": ore, "Organismo Pre-esistente": "",
            ORG: org, "Preferenza Soddisfatta": "1ª", "Data Attivazione": "",
            "Status": "OK", "Note": "",
        }

    # due cooperative con lo STESSO prefisso lungo: i fogli si troncano uguali
    # e vengono deduplicati -> il vecchio ritrovamento per prefisso sbagliava
    org_a = "CONSORZIO COOPERATIVE SOCIALI ALFA"
    org_b = "CONSORZIO COOPERATIVE SOCIALI BETA"
    df = pd.DataFrame([
        _riga_report("1", org_a, 50),
        _riga_report("2", org_a, 40),
        _riga_report("3", org_b, 30),
    ])
    riep = app.costruisci_riepilogo_gruppo(df, 45)
    crit = app.costruisci_criticita(df)
    wb = load_workbook(io.BytesIO(app.genera_excel(df, riep, crit)))
    std = ("Assegnazioni", "Riepilogo Gruppo", "Riepilogo Economico", "Criticita")

    def _norm(s):
        return app.normalizza_nome(s)

    errori = 0
    for s in wb.sheetnames:
        if s in std:
            continue
        ws = wb[s]
        titolo = ws.cell(row=2, column=1).value  # nome org nell'intestazione
        hdr = [ws.cell(row=6, column=c).value for c in range(1, ws.max_column + 1)]
        ci_cod = hdr.index("Codice Iscrizione") + 1
        # a quale org appartengono i codici presenti nel foglio?
        codici = []
        for rr in range(7, ws.max_row + 1):
            v = ws.cell(row=rr, column=ci_cod).value
            if v is not None and not str(v).startswith("TOTALE"):
                codici.append(str(v))
        org_dati = {df.loc[df["Codice Iscrizione"] == c, ORG].iloc[0]
                    for c in codici if (df["Codice Iscrizione"] == c).any()}
        # il titolo del foglio deve coincidere con l'organismo dei dati
        if len(org_dati) == 1 and _norm(titolo) != _norm(next(iter(org_dati))):
            errori += 1
    check("fogli per-organismo: titolo coerente con i dati (prefissi uguali)",
          errori == 0, f"{errori} fogli mal attribuiti")

    # ogni organismo ha un proprio foglio distinto
    fogli_org = [s for s in wb.sheetnames if s not in std]
    check("un foglio distinto per ciascuna cooperativa", len(fogli_org) == 2,
          str(fogli_org))


def test_settimane_periodi():
    print("\n=== Periodi 14/21 e riparametrazione per finestra ===")

    # settimane_effettive: pieni per riconferme/inizio anno, ridotti in corso
    anno = 2025
    for nome, d, exp in [
        ("nessuna data -> pieni", None, (14, 21)),
        ("pre-anno (riconferma) -> pieni", datetime.date(2025, 6, 1), (14, 21)),
        ("inizio anno 16/09 -> pieni", datetime.date(2025, 9, 16), (14, 21)),
        ("dopo giugno -> zero", datetime.date(2026, 6, 30), (0, 0)),
    ]:
        got = app.settimane_effettive(d, anno)
        check(f"settimane_effettive: {nome}", got == exp, f"ottenuto {got}")

    # finestra 2 (novembre): set-dic ridotto, gen-giu pieno
    n1, n2 = app.settimane_effettive(datetime.date(2025, 11, 1), anno)
    check("finestra novembre: set-dic ridotto e gen-giu pieno",
          0 < n1 < app.SETTIMANE_SET_DIC and n2 == app.SETTIMANE_GEN_GIU,
          f"({n1}, {n2})")
    # finestra 3 (gennaio inoltrato): set-dic azzerato
    n1b, n2b = app.settimane_effettive(datetime.date(2026, 1, 20), anno)
    check("finestra gennaio: set-dic azzerato, gen-giu ridotto",
          n1b == 0 and 0 < n2b < app.SETTIMANE_GEN_GIU, f"({n1b}, {n2b})")

    # deriva_anno_start
    check("anno_start da '2025/2026'", app.deriva_anno_start({"anno": "2025/2026"}) == 2025)
    check("anno_start da data riferimento nov.",
          app.deriva_anno_start({}, datetime.date(2025, 11, 1)) == 2025)
    check("anno_start assente -> None", app.deriva_anno_start({}, None) is None)

    # riconciliazione economica: annuo == set-dic + gen-giu (al centesimo)
    df = pd.DataFrame({
        "Ore Assegnate": [10, 10, 20],
        "Settimane set-dic": [14, 8, 14],
        "Settimane gen-giu": [21, 21, 21],
        "Organismo Assegnato dall'Algoritmo": ["A", "A", "B"],
    })
    eco = app.calcola_colonne_economiche(df)
    imp_diff = (eco["Imponibile (EUR)"]
                - (eco["Imponibile set-dic (EUR)"] + eco["Imponibile gen-giu (EUR)"])).abs()
    tot_diff = (eco["Totale (EUR)"]
                - (eco["Totale set-dic (EUR)"] + eco["Totale gen-giu (EUR)"])).abs()
    check("imponibile annuo == set-dic + gen-giu", (imp_diff < 0.005).all())
    check("totale annuo == set-dic + gen-giu", (tot_diff < 0.005).all())
    check("settimane totali = set-dic + gen-giu",
          (eco["Settimane totali"] == eco["Settimane set-dic"] + eco["Settimane gen-giu"]).all())
    # riga riparametrata (29 sett.) costa meno della piena (35 sett.), stesse ore
    check("riparametrato (29 sett.) < pieno (35 sett.)",
          eco["Totale (EUR)"].iloc[1] < eco["Totale (EUR)"].iloc[0])

    # end-to-end: colonne settimane in esegui_assegnazione, riparametrate per finestra
    rows = [
        riga("51", "RICONFERMA", "MUNICIPIO ROMA V", "RMIC000051", "IC EPSILON",
             "RMEE000051", "PLESSO EPSILON", "IC", 45, org="COOP A",
             att="2025-09-15 00:00:00"),
        riga("52", "NUOVA ISCRIZIONE", "MUNICIPIO ROMA V", "RMIC000051", "IC EPSILON",
             "RMEE000051", "PLESSO EPSILON", "IC", 20, enti="COOP A\nCOOP B",
             att="2025-11-17 00:00:00"),
    ]
    df_s, cm_s, _ = app.carica_dati(costruisci_mesis(rows))
    res_s = app.esegui_assegnazione(df_s, cm_s, 45, 50, auto_ambito=True,
                                    anno_start=2025)
    by = res_s[0].set_index("Codice Iscrizione")
    check("colonne settimane presenti nel risultato",
          {"Settimane set-dic", "Settimane gen-giu", "Settimane totali"}
          <= set(res_s[0].columns))
    check("51 (settembre) settimane piene = 35",
          by.at["51", "Settimane totali"] == app.SETTIMANE_ANNO,
          str(by.at["51", "Settimane totali"]))
    check("52 (novembre) riparametrato < 35",
          by.at["52", "Settimane totali"] < app.SETTIMANE_ANNO,
          str(by.at["52", "Settimane totali"]))
    # senza anno_start, nessuna riparametrazione: tutti a 35
    res_no = app.esegui_assegnazione(df_s, cm_s, 45, 50, auto_ambito=True)
    check("senza anno_start: nessuna riparametrazione (tutti 35)",
          (res_no[0]["Settimane totali"] == app.SETTIMANE_ANNO).all())


def test_etichetta_gruppo():
    print("\n=== Unit: etichetta_gruppo ===")
    for cod, desc, atteso in [
        ("IC:RMIC8E0001", "I.C. VALENTE", "Istituto Comprensivo I.C. VALENTE (RMIC8E0001)"),
        ("IC:RMIC8E0001", "", "Istituto Comprensivo RMIC8E0001"),
        ("COM:Ambito 10", "", "Infanzia comunale — Ambito 10"),
        ("PAR:Municipio V", "", "Scuola paritaria — Municipio V"),
        ("AMB:12 - Paritario", "", "Ambito 12 - Paritario"),
        ("PLESSO:RM1AX", "", "Plesso RM1AX"),
        ("", "", ""),
    ]:
        got = app.etichetta_gruppo(cod, desc)
        check(f"etichetta_gruppo({cod!r}, {desc!r})", got == atteso, f"ottenuto {got!r}")


if __name__ == "__main__":
    test_unitari()
    test_etichetta_gruppo()
    if os.path.exists(MESIS):
        test_file_reale(MESIS)
    else:
        print(f"\n(SKIP) File MESIS reale non trovato ({MESIS}): "
              "salto i test di integrazione sul file reale.")
    test_sintetici()
    test_vincolo_45h_riassegnazione()
    test_risoluzione_per_plesso()
    test_corso_anno_sintetico()
    test_cronologia()
    test_cronologia_movimento_netto()
    test_attribuzione_report()
    test_settimane_periodi()

    print()
    if FALLITI:
        print(f"### {len(FALLITI)} TEST FALLITI ###")
        for n in FALLITI:
            print(" -", n)
        sys.exit(1)
    print("### TUTTI I TEST PASSANO ###")
